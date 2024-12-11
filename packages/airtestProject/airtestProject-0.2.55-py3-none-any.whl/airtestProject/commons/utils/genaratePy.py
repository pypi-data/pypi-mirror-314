"""
根据Excel表格生成python测试用例
"""
import os
import openpyxl

from airtestProject import config
from airtestProject.commons.utils.excelPicture import get_id_name
from utils import *


# Utils = Utils()


def get_execls(project):
    """
    获取Excel文件
    return test_case_path_list
    """
    test_case_path = os.path.join(config.ROOT_PATH, "cases", project)
    test_case_path_list = []
    for root, pycache, files in os.walk(test_case_path):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                test_case_path_name = root + '\\' + file
                test_case_path_list.append(test_case_path_name)
    print(test_case_path_list)
    return test_case_path_list


def parse_excel(project="zg", page=0):
    pathList = get_execls(project)
    test_case_data = {}
    test_case_dic = {}
    if pathList is None:
        return
    for file in pathList:
        file_name = file.split('\\')[-1].split('.')[0]
        wb = openpyxl.load_workbook(file)
        sheet = wb.worksheets[page]
        result_dict = {}
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row=row, column=1).value
            if key in result_dict:
                # 如果已经存在，将当前行号添加到对应的list中
                result_dict[key].append(row)
            else:
                if key is None:
                    continue
                # 如果不存在，创建一个新的list，并将行号添加进去
                result_dict[key] = [row]
        for key, vals in result_dict.items():
            operate_dic = {}
            for val in vals:
                operate_key = sheet.cell(val, 2).value
                operate_value = sheet.cell(val, 3).value
                if "_xlfn.DISPIMG" in str(operate_value):
                    operate_dic[operate_key] = [operate_value[1::]]
                elif ',' in str(operate_value):
                    value = operate_value.split(',')
                    operate_dic[operate_key] = value
                else:
                    operate_dic[operate_key] = [operate_value]
            test_case_data[key] = operate_dic
        test_case_dic[file_name] = test_case_data

        ## 把图片id替换为图片路径
        id_name = get_id_name(project, file, page)
        for key, value in test_case_dic[file_name].items():
            for k, v in value.items():
                if v[0] in id_name.keys():
                    test_case_dic[file_name][key][k][0] = id_name[v[0]]


    print(test_case_dic)
    return test_case_dic


if __name__ == '__main__':
    parse_excel("zhange", 0)

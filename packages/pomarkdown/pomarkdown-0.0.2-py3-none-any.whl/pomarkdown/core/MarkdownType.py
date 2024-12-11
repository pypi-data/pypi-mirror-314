import base64
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from pofile import mkdir


class MainMarkdown():

    def markdown_link_image_to_base64(self, markdown_path):
        markdown_path = pathlib.Path(markdown_path)
        if markdown_path.is_file() and markdown_path.suffix == ".md":
            i = 0  # 图片计数变量
            image_base64_list = []  # 图片base64编码列表
            markdown_file = open(markdown_path, encoding="utf-8")
            new_markdown_file = open(markdown_path.parent.joinpath(markdown_path.name[:-3:] + "(base64).md"), "x",
                                     encoding="utf-8")
            markdown_file = markdown_file.readlines()
            for line in markdown_file:
                # 按行遍历文件
                if line[:2:] == "![":
                    # 如果该行有链接图片
                    i = i + 1
                    this_image_path = pathlib.Path(line[line.rfind("(") + 1:line.rfind(")"):])
                    if not this_image_path.is_absolute():
                        # Python文件和markdown文件通常不在同一路径，相对路径基准不同，故需要判断链接图片路径是否为绝对路径。
                        this_image_path = markdown_path.parent.joinpath(this_image_path).resolve()
                    this_image_file = open(this_image_path, "rb")
                    this_image_base64_data = base64.b64encode(this_image_file.read())
                    # 对链接的图片进行base64编码
                    line = line[:line.index("("):] + f"[image{str(i)}]" + "\n"
                    image_base64_list.append(
                        f"[image{str(i)}]: " + "data:image/" + this_image_path.suffix + ";base64," + str(
                            this_image_base64_data)[2:-1:] + "\n")
                new_markdown_file.writelines(line)
            for i in image_base64_list:
                # 遍历存放base64编码的列表
                if markdown_file[-1] != "\n":
                    new_markdown_file.writelines("\n")
                # 判断最后一行是否为空行，如果不是则写入换行符
                new_markdown_file.writelines(i)
                # 将图片的base64编码写入新文件
        elif not markdown_path.is_file():
            print("markdown文件路径输入有误，请检查！")
        elif markdown_path.suffix != ".md":
            print("请输入.md文件的路径！")

    def check_local_dir_image_link_markdown(self, markdown_path, image_path):
        """
        :param markdown_path: markdown文件路径
        :param image_path: 本地图片存放路径
        """
        markdown_path = pathlib.Path(markdown_path)
        image_path = pathlib.Path(image_path)
        # 转换路径为对象
        if markdown_path.is_file() and image_path.is_dir():
            print("检查中……")
            markdown_image_list = []  # markdown文件中链接的图片列表
            markdown_file = open(markdown_path, encoding="utf-8")
            markdown_file = markdown_file.readlines()
            for line in markdown_file:
                # 按行遍历文件
                if line[:2:] == "![":
                    # 如果该行有链接图片
                    this_image_path = pathlib.Path(line[line.rfind("(") + 1:line.rfind(")"):])
                    # 提取链接图片路径
                    if this_image_path.is_absolute():
                        # Python文件和markdown文件通常不在同一路径，相对路径基准不同，故需要判断链接图片路径是否为绝对路径。如果是则直接添加进列表，如果不是则要拼接上markdown所在的路径
                        markdown_image_list.append(this_image_path)
                    else:
                        this_image_path = markdown_path.parent.joinpath(this_image_path).resolve()
                        markdown_image_list.append(this_image_path)
            local_dir_image_list = []
            for i in list(image_path.glob("**/*")):
                local_dir_image_list.append(i.resolve())
            markdown_image_set = set(markdown_image_list)
            local_dir_image_set = set(local_dir_image_list)
            if local_dir_image_set ^ markdown_image_set != (
                    local_dir_image_set ^ markdown_image_set) & local_dir_image_set:
                """
                如果有文件夹中有图片未被链接到markdown中或markdown中链接了其他路径的图片，则local_dir_image_set ^ markdown_image_set不为空
                (local_dir_image_set ^ markdown_image_set) & local_dir_image_set为未被链接到markdown中的图片
                如果两个集合不相等，则代表markdown中链接了其他路径的图片
                """
                print("文档中有链接其他路径的图片，请检查！ ")
                other_dir_image_list = list((local_dir_image_set ^ markdown_image_set) & markdown_image_set)
                print("其他路径图片列表：")
                for i in other_dir_image_list:
                    print(str(i))
                print()
            if (local_dir_image_set ^ markdown_image_set) & local_dir_image_set != set():
                print("未被链接图片列表：")
                for i in local_dir_image_set ^ markdown_image_set & local_dir_image_set:
                    # 打印未被链接图片名列表
                    print(i)
            else:
                print(f"{image_path.resolve()}内无未被链接的图片")
        elif not markdown_path.is_file():
            print("markdown文件路径输入有误，请检查！")
        elif not image_path.is_dir():
            print("图片存放路径输入有误，请检查！")

    def excel2markdown(self, input_file, output_file, sheet_name=None):

        # 创建输出文件夹（如果不存在）
        mkdir(Path(output_file).parent)

        # 加载Excel工作簿
        wb = load_workbook(input_file)

        with open(output_file, 'w', encoding='utf-8') as md_file:
            md_file.write('\n')
        for temp_name in wb.sheetnames:
            if sheet_name is not None and sheet_name != temp_name:
                continue
            # 读取Excel表格
            df = pd.read_excel(input_file, sheet_name=temp_name, engine='openpyxl')
            ws = wb[temp_name]

            # 获取合并单元格信息
            merged_cells = ws.merged_cells.ranges
            merged_values = {}

            # 处理合并单元格
            for merge in merged_cells:
                cell_value = ws.cell(row=merge.min_row, column=merge.min_col).value
                for row in range(merge.min_row, merge.max_row + 1):
                    for col in range(merge.min_col, merge.max_col + 1):
                        merged_values[(row, col)] = cell_value

            # 写入Markdown内容
            with open(output_file, 'a+', encoding='utf-8') as md_file:
                md_file.write(f"\n")
                md_file.write(f"## {temp_name}\n\n")
                md_file.write("| " + " | ".join(df.columns) + " |\n")
                head = " | ".join(["---"] * len(df.columns))
                md_file.write("|  " + head + " |\n")

                for index, row in df.iterrows():
                    formatted_data = []
                    for col_index, col in enumerate(df.columns):
                        # 跳过以 "Unnamed" 开头的列
                        if col.startswith("Unnamed"):
                            continue
                        # 获取当前单元格的值，如果它是合并单元格的一部分，则使用合并单元格的值
                        cell_value = merged_values.get((index + 1, col_index + 1), row[col])
                        formatted_data.append(str(cell_value))

                    md_file.write("| " + " | ".join(formatted_data) + " |\n")

        print("all sheets has been converted to markdown successfully")

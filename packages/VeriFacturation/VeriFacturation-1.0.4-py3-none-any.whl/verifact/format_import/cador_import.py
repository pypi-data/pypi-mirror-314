from openpyxl import load_workbook
from PySide6.QtWidgets import QMessageBox
from pathlib import Path
import datetime
import polars as pl

def test_date(date):
        """Test si une variable est sous un format de date"""
        return isinstance(date, datetime.datetime)
    
def extract_date(jour_facture, date_journal):
        """Renvoie la date d'une écriture d'un journal Cador
        Args:
            jour_facture (str) : jour indiqué pour la ligne d'écriture
            date_journal (datetime) : date indiquée pour le mois du journal
        Returns:
            ma_date (datetime) : date de l'écriture
        """

        jour = int(jour_facture)
        mois = int(date_journal.month)
        annee = int(date_journal.year)
        ma_date = datetime.date(annee, mois, jour)
        #format_date = ma_date.strftime("%d/%m/%Y")
        return ma_date
    
def validate_account(facture: dict, key: str):
    """Vérifie si la facture est valide"""
    return (isinstance(facture["CompteNum"], str) and 
            facture["CompteNum"].startswith(key) and
            isinstance(facture["Debit"], (int, float)) and 
            isinstance(facture["Credit"], (int, float)))
    
def import_log_cador(filename, key):
    """Liste les factures de ventes du journal de vente de Cador"""
    if Path(filename).suffix.lower() != ".xlsx":
        QMessageBox.critical(None, "Erreur de format", 
        "Le format CADOR nécessite un fichier .xlsx")
        return
    
    wb = load_workbook(filename=filename)
    ws = wb.worksheets[0]
    
    # Vérifie que le type de fichier est correct
    if ws.cell(1, 1).value != 'Edition Journaux':
        QMessageBox.critical(None, "Erreur",
            "Il ne s'agit pas d'un journal de vente")
        return
    
    # Nombre de lignes dans la feuille
    nb_lignes = ws.max_row
    liste_factures = []
    date_journal = None
    
    # Récupère la liste des factures de ventes
    for ligne in range(1, nb_lignes + 1):
        if test_date(ws.cell(ligne, 5).value):
            date_journal = ws.cell(ligne, 5).value
        
        try:
            facture = {
                "PieceRef": str(ws.cell(ligne, 2).value),
                "EcritureDate": extract_date(ws.cell(ligne, 1).value, date_journal),
                "EcritureLib": ws.cell(ligne, 3).value,
                "CompteNum": str(ws.cell(ligne, 4).value),
                "Debit": float(str(ws.cell(ligne, 6).value)),
                "Credit": float(str(ws.cell(ligne, 7).value))
            }
        except Exception as e:
            continue
        
        # Ajoute la facture à la liste si celle-ci est valide
        if validate_account(facture, key):
            liste_factures.append(facture.copy())
    
    # Tri ma liste de factures par numéro de pièce
    liste_factures.sort(key=lambda x: x["PieceRef"])
    
    # Fermer le classeur Excel
    wb.close()
    
    # Transforme la liste de factures en dataframe
    df = pl.DataFrame(liste_factures, orient="row")
    
    # Renvoyer le dataframe des factures
    return df
""" itu_appendix42 """

# Table of International Call Sign Series (Appendix 42 to the RR)

# Based on this page ...
#     https://www.itu.int/en/ITU-R/terrestrial/fmd/Pages/glad.aspx
# Visit the following page ...
#     https://www.itu.int/en/ITU-R/terrestrial/fmd/Pages/call_sign_series.aspx
# Download via the .xlsx button and produce a file like this ...
#     CallSignSeriesRanges-998049b7-c007-4e71-bac6-d2393eaa83ef.xlsx
#     CallSignSeriesRanges-c3ce6efb-d36c-4e44-8fff-083b4aab1c09.xlsx
# The following code looks for the newest file of that name pattern in your Download's directory.
# Under windows that's C:\Users\YourUsername\Downloads\, under Linux or MacOS it's ~/Downloads

import sys
import re
from os.path import join, expanduser, getmtime
from glob import glob
from string import ascii_uppercase, digits

from openpyxl import load_workbook

from itu_appendix42.iso3661_mapping_from_itu import iso3661_mapping_from_itu

class ItuAppendix42():
    """ ItuAppendix42 """

    DOWNLOAD_FOLDER = 'Downloads'
    FILENAME_PATTERN = 'CallSignSeriesRanges-*-*-*-*.xlsx'

    _forward = None
    _reverse = None
    _regex = None
    _regex_c = None

    def __init__(self):
        """ __init__ """

        if not self.__class__._forward:
            ws = self._find_worksheet()
            self._build_forward(ws)
            # Further processing reduces this data using regex definition methods
            self._optimize_duplicates()
            # Finally, we build a reverse maping, should we need it
            self._build_reverse()
        if not self.__class__._regex_c:
            self._build_regex()
            self.__class__._regex_c = re.compile(self.__class__._regex, re.ASCII|re.IGNORECASE)

    def dump(self):
        """ dump """
        for k in sorted(self.__class__._forward):
            callsign = k
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            country = self.__class__._forward[k]
            print('%-10s : %s' % (callsign, country))

    def rdump(self):
        """ rdump """
        for k in sorted(self.__class__._reverse):
            print('%-70s : %s' % (k, ','.join(self.__class__._reverse[k])))

    def match(self, line):
        """ match """
        return self.__class__._regex_c.match(line.upper())

    def fullmatch(self, line):
        """ match """
        return self.__class__._regex_c.fullmatch(line.upper())

    def findall(self, line):
        """ findall """
        return [''.join(v) for v in self.__class__._regex_c.findall(line.upper())]

    def _find_worksheet(self):
        """ _find_worksheet """
        try:
            filename = self._find_filename()
        except:
            raise FileNotFoundError(self.__class__.FILENAME_PATTERN) from None
        wb = load_workbook(filename=filename, data_only=True)
        if 'Exported data' != wb.sheetnames[0]:
            raise FileNotFoundError(filename)
        #ws = wb.active
        ws = wb['Exported data']
        return ws

    def _find_filename(self):
        """ _find_filename """
        dirname = join(expanduser('~'), self.__class__.DOWNLOAD_FOLDER)
        a = []
        #
        # Changed in version 3.10: Added the root_dir and dir_fd parameters.
        # for filename in glob(self.__class__.FILENAME_PATTERN, root_dir=dirname):
        #
        for filename in glob(dirname + '/' + self.__class__.FILENAME_PATTERN):
            mtime = getmtime(join(dirname, filename))
            a.append((filename, mtime))
        l = sorted(a, key=lambda item: item[1])
        a = l[-1]
        filename = join(dirname, a[0])
        return filename

    def _build_forward(self, ws):
        """ _build_forward """
        self.__class__._forward = {}
        for v in list(ws.values)[1:]:
            callsign = v[0]
            country = v[1]
            callsign = self._optimize_callsign(callsign)
            if country in iso3661_mapping_from_itu:
                country = iso3661_mapping_from_itu[country]['iso3661'] + '/' + iso3661_mapping_from_itu[country]['iso3661_name']
            self.__class__._forward[callsign] = country

    def _optimize_callsign(self, callsign):
        """ _optimize_callsign """
        # ['5XA - 5XZ']
        callsign_low, callsign_high = callsign.split(' - ')
        # each is three char's long

        if callsign_low[2] == 'A' and callsign_high[2] == 'Z':
            if callsign_low[0:2] == callsign_high[0:2]:
                return callsign_low[0:2] + '[A-Z]'
            if callsign_low[1] == 'A' and callsign_high[1] == 'Z' and callsign_low[0:1] == callsign_high[0:1]:
                return callsign_low[0:1] + '[A-Z][A-Z]'
            if callsign_low[1] == '0' and callsign_high[1] == '9' and callsign_low[0:1] == callsign_high[0:1]:
                return callsign_low[0:1] + '[0-9][A-Z]'

        # For Egypt, Fiji, etc there could be an A-M & N-Z split on the third letter!
        if callsign_low[2] == 'A' and callsign_high[2] == 'M' and callsign_low[0:2] == callsign_high[0:2]:
            if callsign_low[0:2] == callsign_high[0:2]:
                return callsign_low[0:2] + '[A-M]'

        if callsign_low[2] == 'N' and callsign_high[2] == 'Z' and callsign_low[0:2] == callsign_high[0:2]:
            if callsign_low[0:2] == callsign_high[0:2]:
                return callsign_low[0:2] + '[N-Z]'

        return callsign

    def _optimize_duplicates(self):
        """ _optimize_duplicates """

        def dedup(letter_1, letter_2_begin, letter_2_end, present_country):
            """ dedup """
            letter_2 = letter_2_begin
            while letter_2 <= letter_2_end:
                callsign = letter_1 + letter_2 + '[A-Z]'
                del self.__class__._forward[callsign]
                letter_2 = chr(ord(letter_2) + 1)
            if letter_2_begin == letter_2_end:
                letter_range = letter_2_begin
            else:
                letter_range = '[%s-%s]' % (letter_2_begin, letter_2_end)
            callsign = letter_1 + letter_range + '[A-Z]'
            self.__class__._forward[callsign] = present_country

        # now look for second letter sequences
        for letter_1 in sorted(set([v[0:1] for v in self.__class__._forward])):
            for seq in [digits, ascii_uppercase]:
                present_country = None
                letter_2_begin = None
                letter_2_end = None
                for letter_2 in seq:
                    callsign = letter_1 + letter_2 + '[A-Z]'
                    if callsign not in self.__class__._forward:
                        # quite common - this is a non allocated letter sequence
                        if present_country:
                            dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                        present_country = None
                        letter_2_begin = None
                        letter_2_end = None
                        continue
                    if not present_country:
                        # first find of country
                        present_country = self.__class__._forward[callsign]
                        letter_2_begin = letter_2
                        letter_2_end = letter_2
                        continue
                    if present_country != self.__class__._forward[callsign]:
                        # changed country
                        dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                        present_country = self.__class__._forward[callsign]
                        letter_2_begin = letter_2
                        letter_2_end = letter_2
                        continue
                    # continuing country
                    letter_2_end = letter_2

                if present_country:
                    dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                    present_country = None
                    letter_2_begin = None
                    letter_2_end = None

    def _build_reverse(self):
        """_build_reverse """
        self.__class__._reverse = {}
        for k in sorted(self.__class__._forward):
            callsign = k
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            country = self.__class__._forward[k]
            if country not in self.__class__._reverse:
                self.__class__._reverse[country] = []
            self.__class__._reverse[country].append(callsign)

    def _build_regex(self):
        """ _build_regex """

        def expand(s):
            """ expand """
            if len(s) != 3:
                return s
            begin = s[0]
            end = s[2]
            s = ''
            c = begin
            while c <= end:
                s += c
                c = chr(ord(c) + 1)
            return s

        one_letter = '[' + ''.join(sorted([v[0:1] for v in self.__class__._forward if v[-10:] == '[A-Z][A-Z]'])) + ']'

        two_letters = []
        twos = sorted([v[0:2] for v in self.__class__._forward if v[-10:] != '[A-Z][A-Z]' and v[-5:] in ['[A-Z]', '[A-M]', '[N-Z]']])
        for letter_1 in sorted(set([v[0:1] for v in twos])):
            step1 = sorted([v[1:-5] for v in self.__class__._forward if v[0] == letter_1 and v[-5:] == '[A-Z]'])
            step2 = [v[0] for v in step1 if len(v) == 1]
            step3 = [expand(v[1:4]) for v in step1 if len(v) != 1]
            step4 = ''.join(sorted(step2 + step3))
            # The following is in a specific order
            # While this could (and should) be code, there's only some very specific patterns in-use presently
            step4 = step4.replace('ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'A-Z')
            step4 = step4.replace('ABCDEFGHIJKLMNOPQRSTUVWXY', 'A-Y')
            step4 = step4.replace('ABCDEFGHIJKLMNOPQRTUVWXYZ', 'A-RT-Z')
            step4 = step4.replace('ABCDEFGHIJKLMOPQRSTUVWXYZ', 'A-MO-Z')
            step4 = step4.replace('ABCEFGHIJKLMNOPQRSTUVWXYZ', 'A-CE-Z')
            step4 = step4.replace('23456789', '2-9')
            step4 = step4.replace('2345678', '2-8')
            step4 = step4.replace('234567', '2-7')
            step4 = step4.replace('2346789', '2-46-9')
            step4 = step4.replace('2356789', '2-35-9')
            two_letter = letter_1 + '[' + step4 + ']'
            two_letters.append(two_letter)

        three_letters = sorted([v for v in self.__class__._forward if v[-10:] != '[A-Z][A-Z]' and v[-5:] not in ['[A-Z]', '[A-M]', '[N-Z]']])

        prefix_letters = [one_letter + '[A-Z]{0,2}'] + [v + '[A-Z]{0,1}' for v in two_letters] + three_letters
        suffix_regex = '[0-9][0-9A-Z]{0,3}[A-Z]'
        self.__class__._regex = '(' + '|'.join(prefix_letters) + ')(' + suffix_regex + ')'

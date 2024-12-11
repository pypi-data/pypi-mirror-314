"""
_utils/_loader/xlsx.py
========================

.. module:: xlsx
   :platform: Unix
   :synopsis: Provides functions to read and write .xlsx files.

Module Overview
---------------

This module includes functions for reading from and writing to .xlsx files, facilitating the
import and export of spectral data organized in a DataCube format.

Functions
---------

.. autofunction:: _read_xlsx
.. autofunction:: _write_xlsx

"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ._helper import to_cube
from ..._core import DataCube


def _read_xlsx(filepath: str) -> DataCube:
    """
    Read a .xlsx file and convert its contents into a DataCube.

    This function extracts spectral data from the specified Excel file,
    organizing it into a structured DataCube format. It expects the first
    row to contain headers, with 'x' and 'y' values in the respective columns.

    :param filepath: The path to the .xlsx file to be read.
    :type filepath: str
    :return: A DataCube containing the parsed data from the Excel file.
    :rtype: DataCube

    :raises FileNotFoundError: If the specified file does not exist.
    :raises ValueError: If the data cannot be parsed correctly.

    :Example:

    >>> dc = _read_xlsx('path/to/file.xlsx')
    >>> print(dc.shape)  # Output: shape of the DataCube
    """
    data_list = []
    max_x = float('-inf')
    max_y = float('-inf')
    
    # Load the workbook and select the first sheet
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active  # Get the first sheet
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    
    # Iterate over rows and create dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_list.append(row_dict)
        
        # Update max_x and max_y
        x_value = row_dict.get('x', None)
        y_value = row_dict.get('y', None)
        
        if x_value is not None:
            max_x = max(max_x, x_value)
        if y_value is not None:
            max_y = max(max_y, y_value)
    
    cube = to_cube(data_list, len_x=max_x, len_y=max_y)
    
    return DataCube(cube, wavelengths=headers[2:])


def _write_xlsx(datacube: np.ndarray, wavelengths: np.ndarray, filename: str) -> None:
    """
    Write a DataCube to a .xlsx file.

    This function exports the provided DataCube and its associated wavelengths
    to an Excel file.

    :param datacube: The data to be written, structured as a 3D NumPy array.
    :type datacube: np.ndarray
    :param wavelengths: The wavelengths corresponding to the spectral data.
    :type wavelengths: np.ndarray
    :param filename: The name of the file to which the data will be saved (without extension).
    :type filename: str

    :raises ValueError: If the dimensions of the datacube and wavelengths do not match.

    :Example:

    >>> _write_xlsx(dc.cube, dc.wavelengths, 'output_file')
    """
    shape = datacube.shape

    # Create a DataFrame to hold the data
    df = pd.DataFrame()

    # Prepare columns
    cols = [str(wavelength) for wavelength in wavelengths]

    idx = []

    for y in range(shape[1]):
        for x in range(shape[0]):
            spec_ = datacube[x, y, :]

            df_tmp = pd.DataFrame(spec_).T
            df = df.append(df_tmp)

            idx.append(f'x:{x}; y:{y}')

    df.columns = cols
    df.insert(0, column='Point', value=idx)
    df = df.set_index('Point')

    # Write the DataFrame to an Excel file
    df.to_excel(f'{filename}.xlsx')

"""
This module contains the PreDoCS main class PreDoCS and a Control class containing keyword parameters.

Detailed description of PreDoCS

Classes
-------
PreDoCS:
    Offers all methods a user needs for a full stress calculation such as build_model, solve and plot
Control:
    stores information as CPACS filename and path, analysis model ('Song', 'Isotropic'), coordinate system options etc.

.. codeauthor:: Edgar Werthen <Edgar.Werthen@dlr.de>
"""
#   Copyright (c): 2024 Deutsches Zentrum fuer Luft- und Raumfahrt (DLR, German Aerospace Center) <www.dlr.de>. All rights reserved.

import dataclasses
import os
from abc import abstractmethod
from dataclasses import dataclass
from distutils.util import strtobool
from typing import Optional, Union, Callable

import numpy as np
import pandas as pd
from OCC.Core.gp import gp_Dir
from OCC.Display.OCCViewer import rgb_color as color
from cpacs_interface.cpacs_interface import CPACSInterface
from openpyxl import Workbook, load_workbook
from scipy.integrate import solve_ivp
from scipy.linalg import lstsq

from PreDoCS.CrossSectionAnalysis.Display import plot_materials, plot_cross_section_element_values
from PreDoCS.CrossSectionAnalysis.Interfaces import ClassicCrossSectionLoadsWithBimoment, ICrossSectionProcessor, \
    TimoschenkoWithRestrainedWarpingStiffness, CrossSectionInertia, TimoschenkoWithRestrainedWarpingDisplacements, \
    IElement, IElementLoadState, TimoschenkoDisplacements, ClassicCrossSectionLoads
from PreDoCS.CrossSectionAnalysis.Processors import ElementLoadState
from PreDoCS.LoadAnalysis.load_processor import LoadProcessor
from PreDoCS.WingAnalysis.BeamFEM import Beam, get_element_type_from_str, BeamElement4NodeWithoutWarping, \
    BeamElement4NodeWithWarping
from PreDoCS.WingAnalysis.CPACS2PreDoCS import CPACS2PreDoCS
from PreDoCS.WingAnalysis.Display import plot_beam_displacements, plot_beam_internal_loads, plot_beam_3d
from PreDoCS.WingAnalysis.PreDoCSCoord import PreDoCSCoord
from PreDoCS.util.Logging import get_module_logger
from PreDoCS.util.geometry import transform_location_m, transform_displacements_vectors, \
    transform_displacements_list, create_transformation_matrix_by_angles
from PreDoCS.util.occ import calc_cross_section_plane
from PreDoCS.util.util import Vector, get_matrix_interpolation_function

log = get_module_logger(__name__)


@dataclass
class PreDoCS_SolverControl:
    """
    This class contains all control information for the use of PreDoCS calculations.

    Attributes
    ----------
    cpacs_file
        cpacs file, e.g. "/example.xml"
    cpacs_path
        path to the cpacs file, e.g. "../data/CPACS"
    wing_idx
        usually in CPACS defined as follows, 0=wing, 1=htp, 2=vtp. Alternative to wing_uid.
    wing_uid
        uID of the wing. Alternative to wing_idx.
    orientation
        - 'load_reference_points': PreDoCS beam axis is equal to the load reference axis defined by dynamic model
                                   reference points in CPACS. The resulting beam is build with the least amount of kinks
                                   as possible. If all reference points lie on a straight line, only one section is
                                   generated (1D beam). -> 1D or 3D beam
        - 'wingspan': PreDoCS beam axis is the CPACS global y axis and the beam length is equal to the wingspan -> 1D beam
        - 'root-tip': PreDoCS beam axis is between (eta=0, xsi=0.5) and (eta=1, xsi=0.5) -> 1D beam
        - '3d-beam': PreDoCS beam axis is defined by the sections from CPACS. If only one section is defined, the
                     resulting beam is a 1D beam. -> 1D or 3D beam
        - 'rear-spar': PreCoCS beam axis is defined along the rear spar of the wing. At least one spar is necessary in
                       the CPACS definition. If there are no kinks in the rear spar, a 1D beam is generated. Otherwise a
                       3D beam. -> 1D or 3D beam
    sweep_wing
        - False: triggers algorithm to find the first and last feasible wing cross section
                 perpendicular to the beam axis
        - True: triggers algorithm to find the first and last feasible beam nodes.
    loads_are_internal_loads: bool (default: False)
        True if the given loads are internal loads of the beam, False if the given loads are external loads of the beam.
        If the loads are internal loads, no beam FEA is necessary, for external loads beam FEA is performed.
    node_placement: int, list(float), str
            Specifies how the beam nodes are placed
            Available options:
            - int: this defines the number of beam nodes and they are placed equally on the z2 axis
            - list(float): Defines the beam nodes on the z2 axis directly
            - str:
                "ribs": Defines the beam nodes at the intersection of the z2-axis with the ribs.
    processor_type
        Cross section processor for the calculation. Available choices:
            - 'Hybrid'
            - 'Jung'
            - 'Song'
            - 'Hardt'
            - 'Isotropic'
    element_length
        Length of each element of the cross section, usually 0.1 - 0.5.
        Overwritten if segment deflection is active
    segment_deflection
        The segment deflection determines the element length based on the segment curvature.
        If not None, the elements are discretised by dividing the segments into elements, that the deflection
        area between the curve to the discreet geometry for a segment is equal to this parameter.
    parallel_process
        - True (parallel cross section computation)
        - False
    engineering_constants_method
        Engineering constants method to obtain isotropic material from orthotropic material:
            - 'with_poisson_effect': see [Schürmann2007]_, p. 226.
            - 'without_poisson_effect':
            - 'wiedemann': see [Wiedemann2007]_, p. 155.
            - 'song': No stress in 1-direction, no strain in the other direction, see [Song].
    hybrid_processor
        Cross section processor used in the hybrid cross section processor (processor_type == 'Hybrid').
        Available choices:
            - 'Jung'
            - 'Song'
    front_spar_uid
        - CPACS UID of the front spar
        - None, the centre wing box selection algorithm is not triggered
    rear_spar_uid
        - CPACS UID of the rear spar
        - None, the centre wing box selection algorithm is not triggered
    calc_element_load_state_functions
        True, if the element load states should be calculated in the post processing as a function
        of the element coordinate.
    calc_element_load_state_mid_element
        True, if the element load states should be calculated in the post processing
         in the middle (contour-wise) of the element. Can not be used together with calc_element_load_state_functions.
    calc_element_load_state_min_max
        True, if the min and max element load states should be calculated in the post processing.
    mold_angle
        The mould angle influences the spar orientation in the profile. According to figure 24 ("struct_angle") of
        [1], the mould angle is a rotation of the rotor plane around the pitch axis. The
        resulting structural reference plane is going from the blade root origin to the blade tip of the blade geometry
        which includes the prebend and twist. After this rotation of the outer geometry the spars are oriented
        vertically in the blade according to the division point given in figure 23 [1].
        [1] IEA Wind TCP Task 37_10MW.pdf
    simplified_airfoils_epsilon
        The epsilon for the Ramer–Douglas–Peucker algorithm to simplify the airfoils. None for no simplification.
    use_lp
        If True, all materials are converted to lamination parameters.
    lp_material
        Lamination parameter material invariants are calculated with the given material.
    lp_skin
        Set this skin to all optimisation regions. If lp_skin is set lp_material is ignored.
    close_open_ends
        True, if open ends of the cross-section are closed, False if they are left open.
        If True, geometry_closing_threshold is ignored.
    geometry_closing_threshold
        If a geometry is given with two open ends that are not more than te_closing_threshold away from each other,
        the geometry is closed.
    cut_components_uids
        Cut cross-section geometry uids between the components with the given names i.e. to simulate a failed bonding.
        None for no cutting.
        Format for web-shell cut: (web_uid, shell_uid)
        Format for shell-shell cut (intersection_index is 0 or 1): (shell1_uid, shell2_uid, intersection_index)
    dtype
        Datatype used for all calculations.
        Should be complex if the complex step method is used to calculate gradients, otherwise float.
        None for determination of the dtype by the lightworks optimisation control or float.
    beam_fem_element_type
        Element type the PreDoCS beam FE model (only used for external loads).
        For details see PreDoCS.WingAnalysis.BeamFEM.get_element_type_from_str.
        None for automatic selection.
    ensure_cross_sections_in_shell
        Ensure that the cross-sections are placed in a way that even the first and the last cross-section plane
        cuts the leading and trailing edge of the wing shell.
        Only used, if node_placement is an integer.
    """
    cpacs_file: str
    cpacs_path: str = ''
    wing_idx: Optional[int] = None
    wing_uid: Optional[str] = None
    orientation: str = 'load_reference_points'
    # sweep_wing: Optional[bool] = False
    x_axis_definition: str = 'new'
    loads_are_internal_loads: bool = False
    node_placement: Union[int, str, list[float]] = 11
    processor_type: str = 'Hybrid'
    element_length: Optional[float] = 0.1
    segment_deflection: Optional[float] = None
    parallel_process: bool = False
    engineering_constants_method: str = 'with_poisson_effect'
    hybrid_processor: str = 'JungWithoutWarping'
    front_spar_uid: Optional[str] = None
    rear_spar_uid: Optional[str] = None
    calc_element_load_state_functions: bool = True
    calc_element_load_state_mid_element: bool = False
    calc_element_load_state_min_max: bool = False
    mold_angle: float = 0.0
    simplified_airfoils_epsilon: Optional[float] = None
    use_lp: bool = False
    lp_material: Optional['Orthotropic'] = None
    lp_skin: Optional['LaminationParameter'] = None
    close_open_ends: bool = True
    geometry_closing_threshold: float = 1e-2
    cut_components_uids: Optional[list[list[Union[str, int]]]] = None
    dtype: Optional[str] = None
    beam_fem_element_type: Optional[str] = None
    ensure_cross_sections_in_shell: bool = False

    def _init_protocol(self):

        # Get all property names
        property_names = [field.name for field in dataclasses.fields(type(self))]

        # Initialise workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "OptimisationSettings"

        # Write calculation number
        d = ws.cell(row=1, column=1, value='Nr.')

        counter = 0
        for prop_name in property_names:
            attribute = getattr(self, prop_name)
            prop_type = type(attribute)
            if prop_type is dict:
                for key in attribute:
                    d = ws.cell(row=1, column=2 + counter, value=key)
                    counter += 1
            elif prop_type is list:
                d = ws.cell(row=1, column=2 + counter, value=prop_name)
                counter += len(attribute)
            else:
                d = ws.cell(row=1, column=2 + counter, value=prop_name)
                counter += 1

        return wb

    def write_to_protocol(self, path, name):

        # Get full path
        path_name = os.path.join(path, name)

        # Load or initialise protocol as excel spreadsheet
        if os.path.exists(path_name):
            wb = load_workbook(path_name)
        else:
            wb = self._init_protocol()

        # Write data to notebook
        ws = wb.active
        max_row_idx = ws.max_row
        max_column_idx = ws.max_column

        # Write calculation Nr.
        d = ws.cell(row=max_row_idx + 1, column=1, value=max_row_idx)

        for column_idx in range(2, max_column_idx + 1):
            header_cell = ws.cell(row=1, column=column_idx)
            property_name = header_cell.value
            if property_name in ['lp_skin', 'lp_material']:
                # Skip LP skin and material
                continue
            if not header_cell.value:
                list_element_counter += 1
                if list_element_counter <= len(attribute):
                    value = attribute[list_element_counter]
                else:
                    value = ' '
            elif hasattr(self, property_name):
                attribute = getattr(self, property_name)
                if type(attribute) is not list:
                    value = getattr(self, property_name)
                else:
                    list_element_counter = 0
                    value = attribute[list_element_counter]

            else:
                property_names = [p for p in dir(type(self)) if isinstance(getattr(type(self), p), property)]
                for prop_name_compare in property_names:
                    attribute = getattr(self, prop_name_compare)
                    if type(attribute) is dict:
                        if property_name in attribute:
                            value = attribute[property_name]
                            break

            d = ws.cell(row=max_row_idx + 1, column=column_idx, value=value)

        # Save the current settings to the spreadsheet.
        wb.save(path_name)

    def read_from_protocol(self, path, name, nr=None):

        # Get full path
        path_name = os.path.join(path, name)

        # Load protocol
        wb = load_workbook(path_name)
        ws = wb.active
        max_row_idx = ws.max_row
        max_column_idx = ws.max_column

        # Load the calculation specified by nr. or the last calculation
        if nr:
            row_idx = nr + 1
        else:
            row_idx = max_row_idx

        property_name = ''
        for column_idx in range(2, max_column_idx + 1):
            header_cell = ws.cell(row=1, column=column_idx)
            cell = ws.cell(row=row_idx, column=column_idx)

            if header_cell.value:
                property_name = header_cell.value

            if not header_cell.value:
                # Append to list property
                value_list = getattr(self, property_name)
                value_list.append(cell.value)
                setattr(self, property_name, value_list)
            elif hasattr(self, property_name):
                # New property
                attribute = getattr(self, property_name)
                if type(attribute) is not list:
                    setattr(self, property_name, cell.value)
                else:
                    setattr(self, property_name, [cell.value])
            else:
                property_names = [p for p in dir(type(self)) if isinstance(getattr(type(self), p), property)]
                for prop_name_compare in property_names:
                    attribute = getattr(self, prop_name_compare)
                    if type(attribute) is dict:
                        if property_name in attribute:
                            setattr(self, property_name, cell.value)
                            break


class IPreDoCS_Solver:
    @property
    @abstractmethod
    def ctrl(self) -> PreDoCS_SolverControl:
        pass

    @property
    @abstractmethod
    def predocs_coord(self) -> PreDoCSCoord:
        pass

    @property
    @abstractmethod
    def cs_processors(self) -> list[ICrossSectionProcessor]:
        pass

    @property
    @abstractmethod
    def cs_data(self) -> list[tuple[TimoschenkoWithRestrainedWarpingStiffness, CrossSectionInertia]]:
        pass

    @property
    @abstractmethod
    def load_case_names(self) -> list[str]:
        pass

    @property
    @abstractmethod
    def load_processor(self) -> LoadProcessor:
        pass


class IPreDoCSSolverInternal:
    def __init__(self, solver: IPreDoCS_Solver):
        self._solver = solver

    @abstractmethod
    def build_model(self) -> None:
        pass

    @abstractmethod
    def solve(self):
        pass

    @abstractmethod
    def update_structural_model(self) -> None:
        pass

    @property
    @abstractmethod
    def load_case_names(self) -> list[str]:
        pass

    @abstractmethod
    def clear_results(self) -> None:
        pass


class PreDoCSSolverInternalLoads(IPreDoCSSolverInternal):
    def __init__(self, solver: IPreDoCS_Solver):
        super().__init__(solver)

        self._load_case_dict = {}

    def build_model(self):
        assert np.all([lc.internal_loads for lc in self._solver.load_processor.load_cases_imported.values()]), \
            'The SolverControl is set for internal loads, but there are some load cases with external loads.'
        self._load_case_dict = self._solver.load_processor.load_cases_elements

    def solve(self):
        predocs_coord = self._solver.predocs_coord
        cs_processors = self._solver.cs_processors

        # Static analysis for each load case
        cross_section_displacement_dict = {}
        load_states_dict = {}
        min_max_load_states_dict = {}
        for load_case_name, load_case in self._load_case_dict.items():
            cross_section_displacement_list = []
            load_states_list = []
            min_max_load_states_list = []

            forces = load_case.force_vec
            moments = load_case.moment_vec

            for i_cs, (z2_cross_section, cs_processor) in enumerate(zip(predocs_coord.z2_cs, cs_processors)):
                if self._solver.ignore_warping:
                    cs_loads = ClassicCrossSectionLoads(Vector(forces[i_cs, :]), Vector(moments[i_cs, :]))
                else:
                    cs_loads = ClassicCrossSectionLoadsWithBimoment(Vector(forces[i_cs, :]), Vector(moments[i_cs, :]), 0)
                log.debug(f'lc "{load_case_name}"; cs {i_cs}: {forces[i_cs, :]}; {moments[i_cs, :]}')
                element_reference_length_dict = cs_processor.discreet_geometry.element_reference_length_dict

                # Calc cross section displacements
                cross_section_displacements = cs_processor.calc_displacements(cs_loads)
                cross_section_displacement_list.append(cross_section_displacements)

                # Calculate the load states from the cross section displacements
                if self._solver.ctrl.calc_element_load_state_functions:
                    load_state = cs_processor.calc_element_load_states(cross_section_displacements)
                    load_states_list.append(load_state)
                elif self._solver.ctrl.calc_element_load_state_mid_element:
                    load_state = cs_processor.calc_element_load_states(cross_section_displacements)
                    load_state_eval = {}
                    for e, ls in load_state.items():
                        # Calc load states at the middle of the element
                        s_analysis = element_reference_length_dict[e] / 2
                        load_state_eval[e] = ElementLoadState(
                            {k: v(s_analysis) for k, v in ls.strain_state.items()},
                            {k: v(s_analysis) for k, v in ls.stress_state.items()},
                        )
                    load_states_list.append(load_state_eval)
                if self._solver.ctrl.calc_element_load_state_min_max:
                    min_max_load_state = cs_processor.calc_element_min_max_load_states(cross_section_displacements)
                    min_max_load_states_list.append(min_max_load_state)

            cross_section_displacement_dict[load_case_name] = cross_section_displacement_list
            load_states_dict[load_case_name] = load_states_list
            min_max_load_states_dict[load_case_name] = min_max_load_states_list

        # Integrate beam displacements
        beam_displacement_function_dict = self._integrate_cs_displacements(cross_section_displacement_dict)

        return (
            cross_section_displacement_dict,
            beam_displacement_function_dict,
            load_states_dict,
            min_max_load_states_dict
        )

    @staticmethod
    def integrate_cs_displacements(
            predocs_coord,
            cross_section_displacements_func,
            num_min_integration_steps: int = 20,
            dtype=np.float64,
    ) -> 'function(float) -> np.array[6]':
        """
        Integrates the cross section displacements to get the beam displacements, even with internal loads.

        Parameters
        ----------
        num_min_integration_steps
            Minimum number of integration steps along the beam length.

        Returns
        -------
        Beam displacement function along the beam axis.
        """
        z2_cs = predocs_coord.z2_cs

        # Interpolate the transformation matrices between predocs cross-section
        # and cpacs coordinate system along the beam axis
        transformation_aircraft_2_predocs_func = get_matrix_interpolation_function(
            z2_cs, [predocs_coord.transformation_aircraft_2_predocs(z2) for z2 in z2_cs]
        )
        transformation_predocs_2_aircraft_func = get_matrix_interpolation_function(
            z2_cs, [predocs_coord.transformation_predocs_2_aircraft(z2) for z2 in z2_cs]
        )

        # Function of the IVP: dy/dt = ivp_fun(t, y)
        # (t in scipy equals z in PreDoCS,
        # y in scipy is the beam displacement vector in the predocs beam coordinate system)
        # TODO: twisting_derivation is ignored in the IVP
        def ivp_fun(t, y):
            displacements_cpacs = y.flatten()
            # u_cpacs, v_cpacs, w_cpacs, rotx_cpacs, roty_cpacs, rotz_cpacs = displacements_cpacs
            displacements_predocs = transform_displacements_list(
                displacements_cpacs,
                transformation_aircraft_2_predocs_func(t),
            )
            u_predocs, v_predocs, w_predocs, beta_x_predocs, beta_y_predocs, phi_predocs = displacements_predocs
            cs_displacements = cross_section_displacements_func(t)
            gamma_xz, gamma_yz, W_diff, beta_x_diff, beta_y_diff, phi_diff = cs_displacements
            dy_predocs_dt = [
                gamma_xz + beta_y_predocs,  # = U_diff
                gamma_yz - beta_x_predocs,  # = V_diff
                W_diff,
                beta_x_diff,
                beta_y_diff,
                phi_diff,
            ]
            # Transform dy/dt to the CPACS coordinate system
            dy_cpacs_dt = transform_displacements_list(
                dy_predocs_dt,
                transformation_predocs_2_aircraft_func(t),
            )

            return np.array(dy_cpacs_dt, dtype=dtype)

        # Solve IVP
        with np.errstate(under='warn'):
            sol = solve_ivp(
                fun=ivp_fun,
                t_span=[0, predocs_coord.beam_length],
                y0=np.zeros((6,), dtype=dtype),
                method='RK45',
                vectorized=True,
                max_step=predocs_coord.beam_length / num_min_integration_steps,
            )
        assert sol.success, \
            'Integration the beam displacements from the cross section displacements was not successful'
        z2_res = sol.t
        y_res_cpacs = sol.y.T

        # Transform displacements from cpacs to predocs beam coordinate system
        y_res_predocs_beam = [
            transform_displacements_list(y, predocs_coord.transformation_aircraft_2_wing)
            for y in y_res_cpacs
        ]

        # Interpolate the beam displacements from the IVP solution along the beam axis
        beam_displacements_func = get_matrix_interpolation_function(z2_res, y_res_predocs_beam)

        return beam_displacements_func

    def _integrate_cs_displacements(
            self,
            cross_section_displacement_dict,
            num_min_integration_steps: int = 20
    ) -> dict[str, 'function(float) -> np.array[6]']:
        """
        Integrates the cross section displacements to get the beam displacements, even with internal loads.

        Parameters
        ----------
        num_min_integration_steps
            Minimum number of integration steps along the beam length.

        Returns
        -------
        Dict with beam displacement function along the beam axis for each load case.
        """
        solver = self._solver
        predocs_coord = solver.predocs_coord
        dtype = solver.dtype
        z2_cs = predocs_coord.z2_cs

        beam_displacement_function_dict = {}
        # For each load case
        for lc_name in self.load_case_names:

            # Get displacements of each cross section
            cs_displacements_list = np.array(
                [dp.tolist()[0:6] for dp in cross_section_displacement_dict[lc_name]],
                dtype=dtype,
            )

            # Interpolate the cross section displacements along the beam axis
            cs_displacements_func = get_matrix_interpolation_function(z2_cs, cs_displacements_list)#, kind='cubic')

            # Solve IVP
            beam_displacements_func = self.integrate_cs_displacements(
                predocs_coord,
                cs_displacements_func,
                num_min_integration_steps,
                dtype,
            )

            beam_displacement_function_dict[lc_name] = beam_displacements_func

        return beam_displacement_function_dict

    def update_structural_model(self):
        # # Loads have to be updated because of change in pole position
        # self._load_case_dict = self._solver.load_processor.load_cases_elements
        pass

    def clear_results(self) -> None:
        pass

    @property
    def load_case_names(self) -> list[str]:
        return list(self._load_case_dict.keys())


class PreDoCSSolverExternalLoads(IPreDoCSSolverInternal):
    def __init__(self, solver: IPreDoCS_Solver):
        super().__init__(solver)

        if solver.ctrl.beam_fem_element_type is None:
            if solver.ignore_warping:
                self._element_type = BeamElement4NodeWithoutWarping
            else:
                self._element_type = BeamElement4NodeWithWarping
        else:
            self._element_type = get_element_type_from_str(solver.ctrl.beam_fem_element_type)
        self._load_vectors = None
        self._clamped_end = None
        self._beam = None
        self._beam_displacements_dict = None
        self._node_reactions_vector_dict = None
        # self._beam_displacement_function_dict = None
        self._cross_section_internal_loads_function_dict = None

    def build_model(self):
        assert np.all([not lc.internal_loads for lc in self._solver.load_processor.load_cases_imported.values()]), \
            'The SolverControl is set for external loads, but there are some load cases with internal loads.'
        # Beam FE model creation
        beam = Beam.create_beam(
            self._solver.cs_data,
            self._solver.predocs_coord.z2_cs,
            self._solver.predocs_coord.z2_bn,
            self._solver.predocs_coord.z2_2_point_wing,
            element_type=self._element_type,
            dtype=self._solver.dtype,
        )

        # Boundary conditions, can be the same for all beam calculations
        # TODO set boundary conditions method.
        clamped_end = [(0, i, 0) for i in range(6)]

        # Definition of load cases in fe analysis
        load_vectors = {key: beam.get_load_vector(node_loads=node_loads)
                        for key, node_loads in self._solver.load_processor.load_cases_dof.items()}

        self._load_vectors = load_vectors
        self._clamped_end = clamped_end
        self._beam = beam

    def solve(self):
        predocs_coord = self._solver.predocs_coord
        cs_processors = self._solver.cs_processors
        ctrl = self._solver.ctrl
        beam = self._beam
        clamped_end = self._clamped_end
        load_vectors = self._load_vectors

        # CALCULATION
        # Static analysis for each load case
        beam_displacements_dict = {}
        node_reactions_vector_dict = {}
        load_vectors_list = []
        load_key_list = []

        # Transform load vectors dict to lists and initialise result dictionaries
        for load_key in load_vectors.keys():
            load_key_list.append(load_key)
            load_vectors_list.append(load_vectors[load_key])

        # Do the static analysis for all load cases
        beam_displacements, node_reactions_vectors = beam.static_analysis(clamped_end, load_vectors_list)

        # set the result dictionary for each load case
        for load_idx, beam_displacement in enumerate(beam_displacements):
            beam_displacements_dict[load_key_list[load_idx]] = beam_displacement
            node_reactions_vector_dict[load_key_list[load_idx]] = node_reactions_vectors[load_idx]

        self._beam_displacements_dict = beam_displacements_dict
        self._node_reactions_vector_dict = node_reactions_vector_dict

        # POST PROCESSING
        # Beam and Element displacement, stress and strain analysis

        beam_displacement_function_dict = {}
        cross_section_displacement_function_dict = {}
        cross_section_internal_loads_function_dict = {}
        cross_section_displacement_dict = {}
        load_states_dict = {}
        min_max_load_states_dict = {}
        for lc_name in self._solver.load_case_names:
            # Function of the beam displacements over the beam axis
            def beam_displacement_function(z2, lc_name=lc_name):
                return beam.post_processing(beam_displacements_dict[lc_name], z2)[0]

            beam_displacement_function_dict[lc_name] = beam_displacement_function

            # Function of the cross section displacements over the beam axis
            def cross_section_displacement_function(z2, lc_name=lc_name):
                return beam.post_processing(beam_displacements_dict[lc_name], z2)[1]

            cross_section_displacement_function_dict[lc_name] = cross_section_displacement_function

            # Function of the cross section internal loads
            def cross_section_internal_loads_function(z2, lc_name=lc_name):
                return beam.post_processing(beam_displacements_dict[lc_name], z2)[2]

            cross_section_internal_loads_function_dict[lc_name] = cross_section_internal_loads_function

            # Stain and stress distribution of each cross section
            load_states_list = []
            min_max_load_states_list = []
            cross_section_displacement_list = []
            for i_cs, (z_cross_section, cs_processor) in enumerate(zip(predocs_coord.z2_cs, cs_processors)):
                # The cross section displacements from the FEM analysis
                cross_section_displacements = cross_section_displacement_function_dict[lc_name](z_cross_section)
                cross_section_displacement_list.append(cross_section_displacements)

                # Calculate the load states from the cross section displacements
                if ctrl.calc_element_load_state_functions or ctrl.calc_element_load_state_mid_element:
                    load_state = cs_processor.calc_element_load_states(cross_section_displacements)
                    load_states_list.append(load_state)
                if ctrl.calc_element_load_state_min_max:
                    min_max_load_state = cs_processor.calc_element_min_max_load_states(cross_section_displacements)
                    min_max_load_states_list.append(min_max_load_state)

            load_states_dict[lc_name] = load_states_list
            min_max_load_states_dict[lc_name] = min_max_load_states_list
            cross_section_displacement_dict[lc_name] = cross_section_displacement_list

        self._cross_section_internal_loads_function_dict = cross_section_internal_loads_function_dict

        return (
            cross_section_displacement_dict,
            beam_displacement_function_dict,
            load_states_dict,
            min_max_load_states_dict
        )

    def update_structural_model(self):
        predocs_coord = self._solver.predocs_coord
        # load_processor = self._solver.load_processor

        # # Loads have to be updated because of change in pole position
        # load_vectors = {}
        # for key in load_processor.load_cases_dof.keys():
        #     load_vectors[key] = self._beam.get_load_vector(node_loads=load_processor.load_cases_dof[key])  # Load
        # self._load_vectors = load_vectors

        # Beam FE model creation
        self._beam = Beam.create_beam(
            self._solver.cs_data,
            predocs_coord.z2_cs,
            predocs_coord.z2_bn,
            predocs_coord.z2_2_point_wing,
            element_type=self._element_type,
            dtype=self._solver.dtype,
        )

    def clear_results(self) -> None:
        self._beam_displacements_dict = {}
        self._node_reactions_vector_dict = {}

    @property
    def load_case_names(self) -> list[str]:
        return list(self._load_vectors.keys())


STRESS_UINTS = {
    'N_zz': 'N/m',
    'N_zs': 'N/m',
    'N_zn': 'N/m',
    'N_sn': 'N/m',
    'M_zz': 'N',
    'M_ss': 'N',
    'M_zs': 'N',
    'sigma_zz': 'Pa',
    'sigma_zs': 'Pa',
}


class PreDoCS_SolverBase(IPreDoCS_Solver):
    """
    A PreDoCS calculation model

    Attributes
    ----------
    _ctrl
    _predocs_coord
    _cs_processors
    _load_processor
    _beam_displacement_function_dict
    _beam_displacements_dict
    _node_reactions_vector_dict
    _cross_section_displacement_dict
    _load_states_dict
    """

    def __init__(self, ctrl: PreDoCS_SolverControl):
        """
        The PreDoCS calculation model is initialised only with an instance of Class control.

        The initialisation does not import any CPACS model. This is done by build model.

        Parameters
        ----------
        ctrl: Control
            Instance of class Control.
        """
        # Control section
        self._ctrl = ctrl

        self._cpacs_interface = None
        self._c2p = None
        self._structural_model = None
        self._predocs_coord = None
        self._cs_processors = None
        self._cs_geometry = None
        self._load_processor = None
        self._internal_solver = None
        self._dtype = np.dtype(ctrl.dtype) if ctrl.dtype is not None else np.float64

        self._structural_model_interface = None
        self._cross_section_displacement_dict = None
        self._beam_displacement_function_dict = None
        self._load_states_dict = None
        self._min_max_load_states_dict = None

    @property
    def ignore_warping(self) -> bool:
        processor = self.ctrl.processor_type
        if processor == 'Hybrid':
            processor = self.ctrl.hybrid_processor
        if processor in ['Song', 'Jung']:
            return False
        elif  processor in ['JungWithoutWarping']:
            return True
        else:
            raise ValueError(f'Unknown processor {processor} for PreDoCS solver.')

    @property
    def len_displacement_vector(self) -> int:
        if self.ignore_warping:
            return 6
        else:
            return 7

    @property
    @abstractmethod
    def _cpacs_interface_class(self) -> type:
        pass

    @property
    @abstractmethod
    def _get_element_stiffness_func(self) -> Callable:
        pass

    def build_model(
            self,
            cpacs_interface: CPACSInterface = None,
            skip_solve: bool = False,
            skip_internal_solver_build_model: bool = False,
            skip_material_stiffness_calculation_and_update_processors: bool = False,
            **kwargs,
    ):
        """
        The method build_model imports a CPACS model and builds up the entire PreDoCS model
        and create references to the structural model.

        A PreDoCS model includes:
            - Cross section processors
            - A load processor containing the load cases for calculation
            - Boundary conditions, for now a clamped end is hard coded at the wing root.
            - The beam model based on the cross section stiffness.

        Parameters
        ----------
        cpacs_interface
            CPACS interface to use. If None, a new CPACS interface is created.
        skip_solve
            Skip the initial solve.
        """
        # Create CPACS interface
        if cpacs_interface is None:
            log.info('Solver: create CPACS interface...')
            self._cpacs_interface = self._cpacs_interface_class(
                self.ctrl.cpacs_path,
                self.ctrl.cpacs_file,
                simplified_airfoils_epsilon=self.ctrl.simplified_airfoils_epsilon,
                **kwargs,
            )
            log.info('Solver: create CPACS interface... done')
        else:
            self._cpacs_interface = cpacs_interface

        # Create CPACS2PreDoCS object
        log.info('Solver: create CPACS2PreDoCS...')
        c2p = CPACS2PreDoCS(
            self._cpacs_interface,
            loads_are_internal_loads=self.ctrl.loads_are_internal_loads,
            get_element_stiffness_func=self._get_element_stiffness_func,
            wing_index=self.ctrl.wing_idx,
            wing_uid=self.ctrl.wing_uid,
            **kwargs,
        )
        log.info('Solver: create CPACS2PreDoCS... done')

        log.info('Solver: read CPACS2PreDoCS...')
        c2p._read_cpacs()  # Read all CPACS/TIGL data
        log.info('Solver: read CPACS2PreDoCS... done')
        self._c2p = c2p

        # Instantiate PreDoCS coordinate System
        predocs_coord = PreDoCSCoord.from_c2p(
            c2p,
            self.ctrl.node_placement,
            orientation=self.ctrl.orientation,
            x_axis_definition=self.ctrl.x_axis_definition,
            mold_angle=self.ctrl.mold_angle,
            ensure_cross_sections_in_shell=self.ctrl.ensure_cross_sections_in_shell,
        )

        # DEBUG PLOTS
        debug_plots = bool(strtobool(os.getenv('SOLVER_DEBUG_PLOTS', '0')))
        if debug_plots:
            from OCC.Display.SimpleGui import init_display
            display, start_display, add_menu, add_function_to_menu = init_display()

            from OCC.Core.TopoDS import topods
            from OCC.Core.BRepBuilderAPI import BRepBuilderAPI_MakeFace

            # Display wing
            from lightworks.scripts.lw_display_cpacs_file import display_wing_cell_geom_cpacs_interface
            display_wing_cell_geom_cpacs_interface(display, c2p.wing, random_color=True, explosion_factor=1)

            # CS plane via z2_cs
            for z2 in predocs_coord.z2_cs:
                # Same definition as in CPACS2PreDoCS
                section_reference_point, section_reference_axis, z2_s_cs = predocs_coord.get_section_parameter(z2)
                cross_section_plane = calc_cross_section_plane(section_reference_point,
                                                               section_reference_axis,
                                                               z2_s_cs)
                cross_section_plane_face = topods.Face(
                    BRepBuilderAPI_MakeFace(cross_section_plane, -1, 1, -1, 1).Face()
                )
                display.DisplayShape(cross_section_plane_face, color=color(0, 0, 1), transparency=0.5, update=True)
                display.DisplayShape(cross_section_plane.Location(), color=color(0, 0, 1))
            display.camera.SetDirection(gp_Dir(0, 0, -1))
            display.camera.SetUp(gp_Dir(-1, 0, 0))
            display.FitAll()
            start_display()

        log.info('Solver: generate cross section_data...')
        # Create the cross section geometries at given spanwise positions of the wing and
        # computes the cross section stiffness and inertia data from the cross section geometries.
        # Only the upper and lower wing shell with the cells and the spars are imported.
        cs_geometry, cs_processors = c2p.generate_cross_section_data(
            predocs_coord,
            self.ctrl.processor_type,  # The cross section processor
            parallel_processing=self.ctrl.parallel_process,  # True for parallel processing
            element_length=self.ctrl.element_length,  # Element length for the cross section geometry discretization
            segment_deflection=self.ctrl.segment_deflection,  # This overwrites element_length
            method=self.ctrl.engineering_constants_method,  # Engineering constants method
            hybrid_processor=self.ctrl.hybrid_processor,  # The hybrid processor type e.g. 'Song' or 'Jung'
            close_open_ends=self.ctrl.close_open_ends,
            geometry_closing_threshold=self.ctrl.geometry_closing_threshold,
            cut_components_uids=self.ctrl.cut_components_uids,
            skip_material_stiffness_calculation_and_update_processors=skip_material_stiffness_calculation_and_update_processors,
            dtype=self._dtype,
        )
        log.info('Solver: generate cross section_data... done')
        # # Save all cross section geometries as pickle
        # with open('cs_geometry_init.p', 'wb') as f:
        #     pickle.dump(cs_geometry, f)
        # with open('cs_processors_init.p', 'wb') as f:
        #     pickle.dump(cs_processors, f)

        # Instantiate load processor
        assert c2p.loadcase_dict is not None and len(
            c2p.loadcase_dict) > 0, 'No load cases found. Maybe wrong value for loads_are_internal_loads?'
        load_processor = LoadProcessor(c2p.loadcase_dict, predocs_coord)

        self._predocs_coord = predocs_coord
        self._cs_processors = cs_processors
        self._cs_geometry = cs_geometry
        self._load_processor = load_processor

        if self._ctrl.loads_are_internal_loads:
            self._internal_solver = PreDoCSSolverInternalLoads(self)
        else:
            self._internal_solver = PreDoCSSolverExternalLoads(self)

        if not skip_internal_solver_build_model:
            self._internal_solver.build_model()

        # Initial Solve
        if not skip_solve:
            self._solve()

    def clear_results(self, keep_displacements: bool = False) -> None:
        """
        Remove all calculation results from the object.
        """
        if not keep_displacements:
            self._cross_section_displacement_dict = None
            self._beam_displacement_function_dict = None
        self._load_states_dict = None
        self._min_max_load_states_dict = None

        self._internal_solver.clear_results()

    def _solve(self):
        """
        The method solve calculates the displacement of the beam and  cross sections and
        the load states (strains and stresses) of the elements of each cross section.
        """
        (
            self._cross_section_displacement_dict,
            self._beam_displacement_function_dict,
            self._load_states_dict,
            self._min_max_load_states_dict
        ) = self._internal_solver.solve()

    @staticmethod
    def _extreme_stress_states(load_state, element_length):
        """
        Returns the maximum stress value of an element.

        The element stress function is evaluated at three discreet points along the element length, at the ends and in
        the middle. The most extreme value, positive or negative, is selected for each line load. A Dictionary
        connecting each line load to this extreme stress is returned.

        load_state: ElementLoadState
            The PreDoCS own stress and strain format.
            Stresses and strains are available as functions of the element length.
        element_length: float
            Length of the element.

        Return
        ------
        strain_dict: dict
            keys: selection of ('epsilon_zz', 'epsilon_ss', 'kappa_zz', 'kappa_zs', 'gamma_zs', 'kappa_ss')
            values: extreme strains
        stress_dict: dict
            keys: selection of ('N_zz', 'N_zs', 'N_zn', 'N_sn', 'M_zz', 'M_zs', 'M_ss')
            values: extreme line loads
        """
        strain_dict = {}
        stress_dict = {}

        for state_key, stress_func in load_state.stress_state.items():
            stress_dict[state_key] = max([stress_func(0),
                                          stress_func(element_length / 2),
                                          stress_func(element_length)], key=abs)

        for state_key, strain_func in load_state.strain_state.items():
            strain_dict[state_key] = max([strain_func(0),
                                          strain_func(element_length / 2),
                                          strain_func(element_length)], key=abs)

        return strain_dict, stress_dict

    @staticmethod
    def _get_transformation_matrix_from_beam_displacements(beam_displacements):
        """
        Returns the transformation matrix of a cross section from given beam displacements
        at this cross section position.

        Parameters
        ----------
        beam_displacements: np.ndarry
            Vector with the beam cross section displacements.

        Returns
        -------
        np.ndarray
            4x4 affine transformation matrix for the cross section.
        """
        # set up the translation vector
        translation = np.asarray(beam_displacements[0:3]).reshape(3, 1)

        # A rotation of the beam axis is not representative for a rotation of the cross section
        # It is therefore neglected for now
        # phi = beam_displacements[3]
        # theta = beam_displacements[4]
        # psi = beam_displacements[5]

        return create_transformation_matrix_by_angles(translation=translation, phi=0, theta=0, psi=0)

    # @staticmethod
    # def get_element_id(cs_id, component_id, element_id) -> str:
    #     """
    #     Returns a unique id for an PreDoCS cross section element.
    #
    #     Parameters
    #     ----------
    #
    #     Return
    #     ------
    #     """
    #     return '{0:03d}{1:03d}{2:03d}'.format(cs_id, component_id, element_id)

    def transform_displacements_predocs_2_cpacs(self, df: pd.DataFrame) -> pd.DataFrame:
        transformation_matrix = self.predocs_coord.transformation_wing_2_aircraft

        def df_row_func(row):
            pos = Vector([row['x'], row['y'], row['z']])
            u = Vector([row['ux'], row['uy'], row['uz']])
            rot = Vector([row['rotx'], row['roty'], row['rotz']])
            pos_new = transform_location_m(transformation_matrix, pos)
            u_new, rot_new = transform_displacements_vectors(u, rot, transformation_matrix)
            return pd.Series({
                'x': pos_new.x,
                'y': pos_new.y,
                'z': pos_new.z,
                'ux': u_new.x,
                'uy': u_new.y,
                'uz': u_new.z,
                'rotx': rot_new.x,
                'roty': rot_new.y,
                'rotz': rot_new.z,
            }, index=row.index)

        return df.apply(df_row_func, axis=1)

    def __getstate__(self):
        """Make pickle possible."""

        # Speed up pickle
        # self.clear_results()

        state = self.__dict__.copy()
        state['_cpacs_interface'] = None

        return state

    def __setstate__(self, state):
        """Make pickle possible."""
        self.__dict__.update(state)

    def plot_material(self, cross_section_idx, **kwargs):
        """
        Plot the cross section including their material distribution
        """
        cs_processors = self.cs_processors

        # Plot the cross sections
        plot_materials(cs_processors[cross_section_idx].discreet_geometry, **kwargs)

    def plot_all_materials(self, **kwargs):
        """
        Plot the material distribution of all cross sections
        """
        for cs_processor in self.cs_processors:
            plot_materials(cs_processor.discreet_geometry, **kwargs)

    def plot_loads(self, selected_load_case, load_set=('import', 'transform', 'interpolate'), **kwargs):
        """
        Plot the different load sets of a specified load case.

        The load sets are:
            'import': The loads as they are imported from CPACS
            'transform': The loads as they are imported from CPACS seen from the PreDoCS coordinate system
            'interpolate': The loads in the PreDoCS coordinate system interpolated in the cross section planes.
            'shift': Loads in the PreDoCS coordinate system, in the cross section planes at the beam nodes.
        """
        load_processor = self.load_processor

        # plot loads
        load_figs = load_processor.plot_load_cases(name=selected_load_case,
                                                   load_set=load_set, **kwargs)

    def plot_load_redistribution(self, selected_load_case, **kwargs):
        self.load_processor.plot_load_redistribution(load_case_key=selected_load_case, **kwargs)

    def plot_beam_stiffness_and_inertia(self, n=100, **kwargs):
        """
        Plot he stiffness distribution along the wing axis.
        """
        cs_data = self.cs_data
        predocs_coord = self.predocs_coord

        # Plots the stiffness and inertia distribution
        Beam.plot_stiffness_and_inertia(cs_data, predocs_coord.z2_cs, predocs_coord.beam_length, n, **kwargs)

    def get_beam_displacements(self, load_case: str, z2_list: list[float] = None) -> pd.DataFrame:
        """
        Returns the beam displacements along the beam axis as pandas DataFrame.

        Parameters
        ----------
        load_case
            Load case for which the displacements are calculated.
        z2_list
            z2 positions where the displacements are calculated.
        """
        predocs_coord = self.predocs_coord
        beam_displacement_function_dict = self.beam_displacement_function_dict

        if z2_list is None:
            z2_list = np.linspace(0, predocs_coord.z2_cs[-1], 100)
        beam_displacements_function = beam_displacement_function_dict[load_case]
        beam_displacements = []
        for z2 in z2_list:
            pos = predocs_coord.z2_2_point_wing(z2)
            u = beam_displacements_function(z2)
            beam_displacements.append(np.concatenate((pos, u)))
        return pd.DataFrame(
            beam_displacements,
            columns=['x', 'y', 'z', 'ux', 'uy', 'uz', 'rotx', 'roty', 'rotz'],
            index=pd.Series(z2_list, name='z2_beam'),
        )

    def plot_beam_displacements(self, load_case: str, **kwargs):
        """
        Plot the displacements along the wing axis.
        """
        beam_displacements = self.get_beam_displacements(load_case)
        plot_beam_displacements(
            {'PreDoCS': np.concatenate((
                np.array([beam_displacements.index]).T,
                beam_displacements.loc[:, ['ux', 'uy', 'uz', 'rotx', 'roty', 'rotz']],
            ), axis=1)}, num_plots=6, **kwargs)

    def plot_cut_loads(self, selected_load_case, **kwargs):
        """
        Plot the cut loads along the wing axis.
        """
        predocs_coord = self.predocs_coord
        z_to_plot = np.linspace(predocs_coord.z2_cs[0], predocs_coord.z2_cs[-1], 100)

        cross_section_internal_loads = []
        if self._ctrl.loads_are_internal_loads:
            # Plot given internal loads
            num_plots = 6
            load_case = self._internal_solver._load_case_dict[selected_load_case]
            forces = load_case.force_vec
            moments = load_case.moment_vec
            for cs_idx, z in enumerate(predocs_coord.z2_cs):
                u = [z] + np.hstack([forces[cs_idx, :], moments[cs_idx, :]]).tolist()
                cross_section_internal_loads.append(u)
        else:
            # Plot beam FE model results
            num_plots = 6
            cross_section_internal_loads_function = \
                self._internal_solver._cross_section_internal_loads_function_dict[selected_load_case]
            for z in z_to_plot:
                u = [z] + cross_section_internal_loads_function(z).tolist()
                cross_section_internal_loads.append(u)

        cross_section_internal_loads = np.asarray(cross_section_internal_loads)
        plot_beam_internal_loads({'PreDoCS': cross_section_internal_loads}, num_plots=num_plots, **kwargs)

    def plot_load_states(self, cross_section_idx, selected_load_case,
                         selected_stress_state=None, selected_strain_state=None,
                         stress_source='functions', **kwargs):
        """
        Plot the element wise load state for a selected load case and a selected load state.

        Parameters
        ----------
        cross_section_idx: int
            index of the cross section to plot
        selected_load_case: str
            UID of the load case for plotting
        selected_stress_state: str (default: None)
            Not None, if stresses should be plotted.
            Options: 'N_zz', 'N_zs', 'N_zn', 'N_sn', 'M_zz', 'M_zs', 'M_ss'
            For stresses (not stress fluxes) use options: 'sigma_zz', 'sigma_zs'
        selected_strain_state: str (default: None)
            Not None, if strains should be plotted.
            Options: 'epsilon_zz', 'kappa_zz', 'kappa_zs', 'gamma_zs', 'kappa_ss'
        stress_source: str (default: 'functions')
            Selects, from which source the stresses are taken. Available are:
                'functions': Stresses from the stress state functions. calc_element_load_state_functions must be True.
                'min': Min stresses from the element stress states. calc_element_load_state_min_max must be True.
                'max': Max stresses from the element stress states. calc_element_load_state_min_max must be True.
        plot_value_numbers: bool
            plot values in the diagram if True
        """
        assert selected_strain_state or selected_stress_state, 'There must be one strain or stress state given.'
        cs_processors = self.cs_processors

        # Plot the stress/strain flow for all cross section elements
        cs_processor = cs_processors[cross_section_idx]
        if stress_source == 'functions':
            load_states = self.load_states_dict[selected_load_case][cross_section_idx]
        elif stress_source == 'min':
            load_states = {e: load_states[0]
                           for e, load_states in
                           self.min_max_load_states_dict[selected_load_case][cross_section_idx].items()}
        elif stress_source == 'max':
            load_states = {e: load_states[1]
                           for e, load_states in
                           self.min_max_load_states_dict[selected_load_case][cross_section_idx].items()}
        else:
            raise RuntimeError(f'Unknown stress_source "{stress_source}"')

        def get_element_load_state(element, load_state, selected_stress_state):
            if selected_stress_state:
                # Stress states
                if selected_stress_state in load_state.stress_state:
                    return load_state.stress_state[selected_stress_state]
                elif selected_stress_state == 'sigma_zz':
                    if stress_source == 'functions':
                        def res_func(s, stress_state=load_state.stress_state['N_zz'], thickness=element.thickness):
                            return stress_state(s) / thickness

                        return res_func
                    else:
                        return load_state.stress_state['N_zz'] / element.thickness
                elif selected_stress_state == 'sigma_zs':
                    if stress_source == 'functions':
                        def res_func(s, stress_state=load_state.stress_state['N_zs'], thickness=element.thickness):
                            return stress_state(s) / thickness

                        return res_func
                    else:
                        return load_state.stress_state['N_zs'] / element.thickness
                else:
                    raise RuntimeError(f'Unknown stress state "{selected_stress_state}"')
            else:
                # Strain states
                return load_state.strain_state[selected_strain_state]

        element_functions = {
            element: get_element_load_state(element, load_state, selected_stress_state)
            for element, load_state in load_states.items()
        }
        plot_cross_section_element_values(
            cross_section=cs_processor,
            value_dict=element_functions,
            values_are_functions=stress_source == 'functions',
            scale_unit=STRESS_UINTS[selected_stress_state] if selected_stress_state and selected_stress_state in STRESS_UINTS else '?',
            **kwargs
        )

    def plot_3d(
            self,
            selected_load_case,
            **kwargs,
    ):
        plot_beam_3d(
            c2p=self.c2p,
            predocs_coord=self.predocs_coord,
            cs_processors=self.cs_processors,
            beam_displacements_function=self.beam_displacement_function_dict[selected_load_case],
            **kwargs,
        )

    @property
    def dtype(self):
        return self._dtype

    @property
    def cpacs_interface(self) -> CPACSInterface:
        return self._cpacs_interface

    @cpacs_interface.setter
    def cpacs_interface(self, value):
        if value is None or (isinstance(value, bool) and value):
            del self._cpacs_interface
            self._cpacs_interface = 'The property cpacs_interface was deleted!'
        elif isinstance(value, CPACSInterface):
            self._cpacs_interface = value
        else:
            raise ValueError(f'Property cpacs_interface can not set with "{value}"')

    @property
    def c2p(self) -> CPACS2PreDoCS:
        return self._c2p

    @c2p.setter
    def c2p(self, value):
        if value is None or (isinstance(value, bool) and value):
            del self._c2p
            self._c2p = 'The property c2p was deleted!'
        elif isinstance(value, CPACS2PreDoCS):
            self._c2p = value
        else:
            raise ValueError(f'Property c2p can not set with "{value}"')

    @property
    def ctrl(self) -> PreDoCS_SolverControl:
        return self._ctrl

    @property
    def predocs_coord(self) -> PreDoCSCoord:
        return self._predocs_coord

    @property
    def cs_processors(self) -> list[ICrossSectionProcessor]:
        return self._cs_processors

    @property
    def cs_data(self) -> list[tuple[TimoschenkoWithRestrainedWarpingStiffness, CrossSectionInertia]]:
        return [(cs_processor.stiffness, cs_processor.inertia) for cs_processor in self._cs_processors]

    @property
    def load_case_names(self) -> list[str]:
        return self._internal_solver.load_case_names

    @property
    def load_processor(self) -> LoadProcessor:
        return self._load_processor

    @property
    def beam_displacement_function_dict(self) -> dict[str, 'function(float) -> np.array[6]']:
        return self._beam_displacement_function_dict

    @property
    def cross_section_displacement_dict(self) -> dict[str, list[TimoschenkoDisplacements | TimoschenkoWithRestrainedWarpingDisplacements]]:
        return self._cross_section_displacement_dict

    @property
    def load_states_dict(self) -> dict[str, list[dict[IElement, ElementLoadState]]]:
        return self._load_states_dict

    @property
    def min_max_load_states_dict(
            self
    ) -> dict[str, list[dict[IElement, tuple[ElementLoadState, ElementLoadState]]]]:
        return self._min_max_load_states_dict

    def _get_sensor_positions_element_mapping(self, sensor_positions: list[Vector]) -> list[tuple[int, IElement]]:
        """
        Returns the PreDoCS cross-section elements for given positions in the global CPACS coordinate system.

        Parameters
        ----------
        sensor_positions
            List of sensor positions in the global CPACS coordinate system.

        Returns
        -------
        List of tuples of cross-section index and cross-section element.
        """
        cs_indices = []
        sensor_positions_cs_3d = []
        sensor_positions_cs = []
        for sensor_position in sensor_positions:
            # Find cross-sections for positions
            z2 = self.predocs_coord.point_2_z2_aircraft(sensor_position)
            cs_dist = [abs(proc.z_beam - z2) for proc in self.cs_processors]
            cs_idx = np.argmin(cs_dist)
            log.debug(f'Sensor position {sensor_position}:')
            log.debug(f'    in cross-section {cs_idx} with z2-distance {cs_dist[cs_idx]:.3f}')
            cs_proc = self.cs_processors[cs_idx]
            z2_cs = cs_proc.z_beam
            sensor_position_cs = transform_location_m(
                self.predocs_coord.transformation_aircraft_2_predocs(z2_cs),
                sensor_position,
            )
            sensor_positions_cs_3d.append(sensor_position_cs)
            log.debug(f'    cross-section z2-distance {sensor_position_cs.z:.3f}')
            sensor_position_cs = Vector([sensor_position_cs.x, sensor_position_cs.y])
            cs_indices.append(cs_idx)
            sensor_positions_cs.append(sensor_position_cs)

        sensor_elements = [None for p in sensor_positions_cs]
        for cs_idx in sorted(set(cs_indices)):
            # Find elements in cross-section
            log.debug(f'Cross-section {cs_idx}:')
            cs_proc = self.cs_processors[cs_idx]
            cs_positions = {i: p for i, (p, i_cs) in enumerate(zip(sensor_positions_cs, cs_indices)) if i_cs == cs_idx}

            distances = []
            for e in cs_proc.discreet_geometry.elements:
                distances.append([e] + [e.position.dist(p) for p in cs_positions.values()])
            df = pd.DataFrame(distances).set_index(0)

            for col, i in zip(df.columns, cs_positions.keys()):
                element = df[col].idxmin()
                element_pos_cs_3d = Vector([element.position.x, element.position.y, 0])
                log.debug(
                    f'    sensor position {sensor_positions[i]} ({sensor_positions_cs_3d[i]}) mapped to element {element.id} @ {element.position}: '
                    f'distance {sensor_positions_cs_3d[i].dist(element_pos_cs_3d):.3f}'
                )
                sensor_elements[i] = element
        assert None not in sensor_elements

        return list(zip(cs_indices, sensor_elements))

    def _cs_displacement_from_list(self, displacements):
        if self.ignore_warping:
            return TimoschenkoDisplacements.from_list(displacements)
        else:
            return TimoschenkoWithRestrainedWarpingDisplacements.from_list(displacements)

    def _get_unity_displacements(self) -> tuple[list[TimoschenkoDisplacements | TimoschenkoWithRestrainedWarpingDisplacements], np.ndarray]:
        """
        Get unity displacements.

        Return
        ------
        list[TimoschenkoWithRestrainedWarpingDisplacements]
            List of unity cross-section displacements for all cross-sections.
        np.ndarray
            Matrix of unity cross-section displacements for all cross-sections.
        """
        len_displacement_vector = self.len_displacement_vector
        zero_displacements = [0] * len_displacement_vector
        unity_load = 1e-6
        unity_displacements = []
        for i in range(len_displacement_vector):
            displacements = zero_displacements.copy()
            displacements[i] = unity_load
            unity_displacements.append(self._cs_displacement_from_list(displacements))
        unity_displacements_matrix = np.eye(len_displacement_vector) * unity_load
        return unity_displacements, unity_displacements_matrix

    def _get_sensor_unity_load_states(
            self,
            sensor_element_mapping: list[tuple[int, IElement]],
            unity_displacements: list[TimoschenkoDisplacements | TimoschenkoWithRestrainedWarpingDisplacements],
    ) -> 'np.ndarray[IElementLoadState]':
        """
        Returns the load states of the elements for the sensor-element-mapping for all unity cross-section-displacements.

        Parameters
        ----------
        sensor_element_mapping
            List of tuples of cross-section index and cross-section element.
        unity_displacements
            List of unity cross-section displacements for all cross-sections.

        Return
        ------
        np.ndarray[IElementLoadState]
            Matrix of unity cross-section displacements for all sensors.
        """
        sensor_unity_load_states = [[None for i in range(len(unity_displacements))] for p in sensor_element_mapping]
        cs_indices = sorted({i for i, e in sensor_element_mapping})
        for i_cs in cs_indices:
            cs_processor = self.cs_processors[i_cs]
            elements = [e for i, e in sensor_element_mapping if i == i_cs]
            for i_displ, cs_displacements in enumerate(unity_displacements):
                element_load_states = cs_processor.calc_element_load_states(cs_displacements)
                for e in elements:
                    i_sensor = sensor_element_mapping.index((i_cs, e))
                    sensor_unity_load_states[i_sensor][i_displ] = element_load_states[e]
        sensor_unity_load_states = np.array(sensor_unity_load_states)
        assert None not in sensor_unity_load_states
        return sensor_unity_load_states

    def _get_cross_section_displacement_recovery_matrices(
            self,
            sensor_element_mapping: list[tuple[int, IElement]],
            sensor_unity_load_states: 'np.ndarray[IElementLoadState]',
            strain_label_list: list[str],
    ) -> tuple[dict[int, np.ndarray], list[int], int]:
        """
        Returns the matrices for the cross-sections that calc the sensor strains from the unity displacements.

        Parameters
        ----------
        sensor_element_mapping
            List of tuples of cross-section index and cross-section element.
        sensor_unity_load_states
            Matrix of unity cross-section displacements for all sensors.
        strain_label_list
            List of strain state labels that are evaluated for each sensor.

        Return
        ------
        dict[int, np.ndarray]
            Dict of cross-section indices to the corresponding cross-section-displacement-strain-recovery-matrix
        list[int]
            Index  of each sensor in the corresponding cross-section.
        int
            Number evaluated of strains per sensor.
        """
        len_displacement_vector = self.len_displacement_vector
        strains_per_sensor = len(strain_label_list)
        cs_indices = sorted({i_cs for i_cs, e in sensor_element_mapping})
        cs_num_sensors = {i_cs: 0 for i_cs in cs_indices}
        sensor_idx_in_cs_list = []
        for i_cs, e in sensor_element_mapping:
            sensor_idx_in_cs_list.append(cs_num_sensors[i_cs])
            cs_num_sensors[i_cs] += 1
        cs_displacement_strain_matrices = {
            i_cs: np.zeros((cs_num_sensors[i_cs] * strains_per_sensor, len_displacement_vector)) for i_cs in cs_indices
        }
        for i_sensor, (i_cs, element) in enumerate(sensor_element_mapping):
            cs_processor = self.cs_processors[i_cs]
            sensor_idx_in_cs = sensor_idx_in_cs_list[i_sensor]
            s_mid = cs_processor.discreet_geometry.element_reference_length_dict[element]
            for i_displ in range(sensor_unity_load_states.shape[1]):
                ls = sensor_unity_load_states[i_sensor, i_displ]
                for i, strain_label in enumerate(strain_label_list):
                    cs_displacement_strain_matrices[i_cs][sensor_idx_in_cs * strains_per_sensor + i, i_displ] = ls.strain_state[strain_label](s_mid)
        return cs_displacement_strain_matrices, sensor_idx_in_cs_list, strains_per_sensor

    def get_cross_section_displacement_from_sensor_strains_setup(
            self,
            sensor_positions: list[Vector],
            strain_label_list: list[str],
    ) -> tuple[
         dict[int, np.ndarray],
         list[tuple[int, IElement]],
         list[int],
         int,
         np.ndarray,
    ]:
        """
        Calculates setup data for the recovery of the cross-section displacements from given strains at the sensors.

        Parameters
        ----------
        sensor_positions
            List of sensor positions in the global CPACS coordinate system.
        strain_label_list
            List of strain state labels that are evaluated for each sensor.

        Return
        ------
        dict[int, np.ndarray]
            Dict of cross-section indices to the corresponding cross-section-displacement-strain-recovery-matrix.
        sensor_element_mapping
            List of tuples of cross-section index and cross-section element.
        list[int]
            Index  of each sensor in the corresponding cross-section.
        int
            Number evaluated of strains per sensor.
        np.ndarray
            Matrix of unity cross-section displacements for all cross-sections.
        """
        sensor_element_mapping = self._get_sensor_positions_element_mapping(sensor_positions)

        unity_displacements, unity_displacements_matrix = self._get_unity_displacements()
        sensor_unity_load_states = self._get_sensor_unity_load_states(sensor_element_mapping, unity_displacements)

        cs_displacement_strain_matrices, sensor_idx_in_cs_list, strains_per_sensor = self._get_cross_section_displacement_recovery_matrices(
            sensor_element_mapping, sensor_unity_load_states, strain_label_list,
        )
        return (
            cs_displacement_strain_matrices,
            sensor_element_mapping,
            sensor_idx_in_cs_list,
            strains_per_sensor,
            unity_displacements_matrix,
        )

    def get_cross_section_displacement_from_sensor_strains(
            self,
            sensor_strains: list[list[float]],
            cs_displacement_strain_matrices: dict[int, np.ndarray],
            sensor_element_mapping: list[tuple[int, IElement]],
            sensor_idx_in_cs_list: list[int],
            strains_per_sensor: int,
            unity_displacements_matrix: np.ndarray,
    ) -> dict[int, TimoschenkoDisplacements | TimoschenkoWithRestrainedWarpingDisplacements]:
        """
        Calculates the cross-section displacements from the given strains at the sensors.

        Parameters
        ----------
        sensor_strains
            List of strains for each sensor.
        cs_displacement_strain_matrices
            Dict of cross-section indices to the corresponding cross-section-displacement-strain-recovery-matrix.
        sensor_element_mapping
            List of tuples of cross-section index and cross-section element.
        sensor_idx_in_cs_list
            Index  of each sensor in the corresponding cross-section.
        strains_per_sensor
            Number evaluated of strains per sensor.
        unity_displacements_matrix
            Matrix of unity cross-section displacements for all cross-sections.

        Return
        ------
        list[TimoschenkoWithRestrainedWarpingDisplacements]
            List of the calculated cross-section displacements for the cross-sections.
        """
        len_displacement_vector = self.len_displacement_vector
        cs_indices = sorted({i_cs for i_cs, e in sensor_element_mapping})
        cs_sensor_strain_dict = {i_cs: {} for i_cs in cs_indices}
        # Assemble cross-section strain vectors
        for i_sensor, ((i_cs, e), sensor_idx_in_cs) in enumerate(zip(sensor_element_mapping, sensor_idx_in_cs_list)):
            for i_strain in range(strains_per_sensor):
                cs_sensor_strain_dict[i_cs][sensor_idx_in_cs * strains_per_sensor + i_strain] = sensor_strains[i_sensor][i_strain]
        cs_strain_vectors = {}
        for i_cs in cs_indices:
            indices = sorted(cs_sensor_strain_dict[i_cs].keys())
            assert indices[0] == 0
            assert indices[-1] == len(indices) - 1
            assert len(set(indices)) == len(indices)
            cs_strain_vector = np.array([cs_sensor_strain_dict[i_cs][i] for i in indices])
            cs_strain_vectors[i_cs] = cs_strain_vector

        # Solve
        cs_displacement_dict = {}
        for i_cs in cs_indices:
            p, res, rnk, s = lstsq(cs_displacement_strain_matrices[i_cs], cs_strain_vectors[i_cs])
            assert rnk == len_displacement_vector
            log.debug(f'lstsq results: residues: {res}, singular values: {s}.')

            cs_displacement_vector = unity_displacements_matrix @ np.array(p)
            cs_displacement_dict[i_cs] = self._cs_displacement_from_list(cs_displacement_vector)

        return cs_displacement_dict

    def get_sensor_load_states(
            self,
            load_case: str,
            sensor_element_mapping: list[tuple[int, IElement]],
            strain_label_list: list[str],
    ) -> list[list[float]]:
        """
        Returns the strains of the sensors for the given load case.

        Parameters
        ----------
        load_case
            Name of the load case.
        sensor_element_mapping
            List of tuples of cross-section index and cross-section element.
        strain_label_list
            List of strain state labels that are evaluated for each sensor.

        Returns
        -------
        list[list[float]]
            List of strains for each sensor.
        """
        load_states_dict = self.load_states_dict[load_case]

        sensor_load_states = []
        for i_cs, element in sensor_element_mapping:
            cs_processor = self.cs_processors[i_cs]
            s_mid = cs_processor.discreet_geometry.element_reference_length_dict[element]
            ls = load_states_dict[i_cs][element]
            sensor_load_states.append([ls.strain_state[strain_label](s_mid) for strain_label in strain_label_list])
        return sensor_load_states

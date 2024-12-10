import csv
import logging
import ntpath
import os
import site
import tempfile
from pathlib import Path


from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from otlmow_converter.DotnotationHelper import DotnotationHelper
from otlmow_converter.OtlmowConverter import OtlmowConverter
from otlmow_model.OtlmowModel.BaseClasses.BooleanField import BooleanField
from otlmow_model.OtlmowModel.BaseClasses.KeuzelijstField import KeuzelijstField
from otlmow_model.OtlmowModel.BaseClasses.OTLObject import dynamic_create_instance_from_uri
from otlmow_model.OtlmowModel.Helpers.generated_lists import get_hardcoded_relation_dict
from otlmow_modelbuilder.OSLOCollector import OSLOCollector
from otlmow_modelbuilder.SQLDataClasses.OSLOClass import OSLOClass

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

enumeration_validation_rules = {
    "valid_uri_and_types": {},
    "valid_regexes": [
        "^https://wegenenverkeer.data.vlaanderen.be/ns/.+"]
}


class SubsetTemplateCreator:
    def __init__(self):
        pass

    @staticmethod
    def _load_collector_from_subset_path(path_to_subset: Path) -> OSLOCollector:
        collector = OSLOCollector(path_to_subset)
        collector.collect_all(include_abstract=True)
        return collector

    def generate_template_from_subset(self, path_to_subset: Path, path_to_template_file_and_extension: Path,
                                      ignore_relations: bool = True, filter_attributes_by_subset: bool = True,
                                      **kwargs):
        tempdir = Path(tempfile.gettempdir()) / 'temp-otlmow'
        if not tempdir.exists():
            os.makedirs(tempdir)
        test = ntpath.basename(path_to_template_file_and_extension)
        temporary_path = Path(tempdir) / test
        instantiated_attributes = (self.generate_basic_template(
            path_to_subset=path_to_subset, temporary_path=temporary_path, ignore_relations=ignore_relations,
            path_to_template_file_and_extension=path_to_template_file_and_extension,
            filter_attributes_by_subset=filter_attributes_by_subset, **kwargs))
        extension = os.path.splitext(path_to_template_file_and_extension)[-1].lower()
        if extension == '.xlsx':
            self.alter_excel_template(path_to_template_file_and_extension=path_to_template_file_and_extension,
                                      temporary_path=temporary_path,
                                      path_to_subset=path_to_subset, instantiated_attributes=instantiated_attributes,
                                      **kwargs)
        elif extension == '.csv':
            self.determine_multiplicity_csv(path_to_template_file_and_extension=path_to_template_file_and_extension,
                                            path_to_subset=path_to_subset,
                                            instantiated_attributes=instantiated_attributes,
                                            temporary_path=temporary_path,
                                            **kwargs)

    def generate_basic_template(self, path_to_subset: Path, path_to_template_file_and_extension: Path,
                                temporary_path: Path, ignore_relations: bool = True, **kwargs):
        list_of_otl_objectUri = None
        if kwargs is not None:
            list_of_otl_objectUri = kwargs.get('list_of_otl_objectUri', None)
        collector = self._load_collector_from_subset_path(path_to_subset=path_to_subset)
        filtered_class_list = self.filters_classes_by_subset(
            collector=collector, list_of_otl_objectUri=list_of_otl_objectUri)
        otl_objects = []
        amount_of_examples = kwargs.get('amount_of_examples', 0)
        model_directory = None
        if kwargs is not None:
            model_directory = kwargs.get('model_directory', None)
        relation_dict = get_hardcoded_relation_dict(model_directory=model_directory)

        generate_dummy_records = 1
        if amount_of_examples > 1:
            generate_dummy_records = amount_of_examples

        for class_object in [cl for cl in filtered_class_list if cl.abstract == 0]:
            if ignore_relations and class_object.objectUri in relation_dict:
                continue
            for _ in range(generate_dummy_records):
                instance = dynamic_create_instance_from_uri(class_object.objectUri, model_directory=model_directory)
                if instance is None:
                    continue
                attributen = collector.find_attributes_by_class(class_object)
                for attribute_object in attributen:
                    attr = getattr(instance, '_' + attribute_object.name)
                    attr.fill_with_dummy_data()
                try:
                    geo_attr = getattr(instance, '_geometry')
                    geo_attr.fill_with_dummy_data()
                except AttributeError:
                    pass
                otl_objects.append(instance)

                DotnotationHelper.clear_list_of_list_attributes(instance)

        converter = OtlmowConverter()
        converter.from_objects_to_file(file_path=temporary_path,
                                          sequence_of_objects=otl_objects, **kwargs)
        path_is_split = kwargs.get('split_per_type', True)
        extension = os.path.splitext(path_to_template_file_and_extension)[-1].lower()
        instantiated_attributes = []
        if path_is_split is False or extension == '.xlsx':
            instantiated_attributes = converter.from_file_to_objects(file_path=temporary_path,
                                                                        path_to_subset=path_to_subset)
        return instantiated_attributes

    @classmethod
    def alter_excel_template(cls, path_to_template_file_and_extension: Path, path_to_subset: Path,
                             instantiated_attributes: list, temporary_path, **kwargs):
        generate_choice_list = kwargs.get('generate_choice_list', False)
        add_geo_artefact = kwargs.get('add_geo_artefact', False)
        add_attribute_info = kwargs.get('add_attribute_info', False)
        highlight_deprecated_attributes = kwargs.get('highlight_deprecated_attributes', False)
        amount_of_examples = kwargs.get('amount_of_examples', 0)
        if add_attribute_info and amount_of_examples == 0:
            amount_of_examples = 1
        wb = load_workbook(temporary_path)
        wb.create_sheet('Keuzelijsten')
        # Volgorde is belangrijk! Eerst rijen verwijderen indien nodig dan choice list toevoegen,
        # staat namelijk vast op de kolom en niet het attribuut in die kolom
        if add_geo_artefact is False:
            cls.remove_geo_artefact_excel(workbook=wb)
        if generate_choice_list:
            cls.add_choice_list_excel(workbook=wb, instantiated_attributes=instantiated_attributes,
                                      path_to_subset=path_to_subset)
        cls.add_mock_data_excel(workbook=wb, rows_of_examples=amount_of_examples)
        if highlight_deprecated_attributes:
            cls.check_for_deprecated_attributes(workbook=wb, instantiated_attributes=instantiated_attributes)
        if add_attribute_info:
            cls.add_attribute_info_excel(workbook=wb, instantiated_attributes=instantiated_attributes)
        cls.design_workbook_excel(workbook=wb)
        wb.save(path_to_template_file_and_extension)
        file_location = os.path.dirname(temporary_path)
        [f.unlink() for f in Path(file_location).glob("*") if f.is_file()]

    def determine_multiplicity_csv(self, path_to_template_file_and_extension: Path, path_to_subset: Path,
                                   instantiated_attributes: list, temporary_path: Path, **kwargs):
        path_is_split = kwargs.get('split_per_type', True)
        if path_is_split is False:
            self.alter_csv_template(path_to_template_file_and_extension=path_to_template_file_and_extension,
                                    temporary_path=temporary_path, path_to_subset=path_to_subset, **kwargs)
        else:
            self.multiple_csv_template(path_to_template_file_and_extension=path_to_template_file_and_extension,
                                       temporary_path=temporary_path,
                                       path_to_subset=path_to_subset, instantiated_attributes=instantiated_attributes,
                                       **kwargs)
        file_location = os.path.dirname(temporary_path)
        [f.unlink() for f in Path(file_location).glob("*") if f.is_file()]

    @classmethod
    def filters_classes_by_subset(cls, collector: OSLOCollector,
                                  list_of_otl_objectUri: [str] = None) -> list[OSLOClass]:
        if list_of_otl_objectUri is None:
            list_of_otl_objectUri = []

        if list_of_otl_objectUri == []:
            return collector.classes
        return [x for x in collector.classes if x.objectUri in list_of_otl_objectUri]

    @staticmethod
    def _try_getting_settings_of_converter() -> Path:
        converter_path = Path(site.getsitepackages()[0]) / 'otlmow_converter'
        return converter_path / 'settings_otlmow_converter.json'

    @classmethod
    def design_workbook_excel(cls, workbook):
        for sheet in workbook:
            dim_holder = DimensionHolder(worksheet=sheet)
            for col in range(sheet.min_column, sheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=25)
            sheet.column_dimensions = dim_holder

    @classmethod
    def add_attribute_info_excel(cls, workbook, instantiated_attributes: list):
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = SubsetTemplateCreator.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            sheet.insert_rows(1)
            for rows in sheet.iter_rows(min_row=2, max_row=2, min_col=1):
                for cell in rows:
                    if cell.value == 'typeURI':
                        value = 'De URI van het object volgens https://www.w3.org/2001/XMLSchema#anyURI .'
                    elif cell.value.find('[DEPRECATED]') != -1:
                        strip = cell.value.split(' ')
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                strip[1])
                        value = dotnotation_attribute.definition
                    else:
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                cell.value)
                        value = dotnotation_attribute.definition
                    newcell = sheet.cell(row=1, column=cell.column, value=value)
                    newcell.alignment = Alignment(wrapText=True, vertical='top')
                    newcell.fill = PatternFill(start_color="808080", end_color="808080",
                                               fill_type="solid")

    @classmethod
    def check_for_deprecated_attributes(cls, workbook, instantiated_attributes: list):
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = SubsetTemplateCreator.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=2):
                for cell in rows:
                    is_deprecated = False
                    if cell.value.count('.') == 1:
                        dot_split = cell.value.split('.')
                        attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                    dot_split[0])

                        if len(attribute.deprecated_version) > 0:
                            is_deprecated = True
                    dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                            cell.value)
                    if len(dotnotation_attribute.deprecated_version) > 0:
                        is_deprecated = True

                    if is_deprecated:
                        cell.value = f'[DEPRECATED] {cell.value}'

    @classmethod
    def find_uri_in_sheet(cls, sheet):
        filter_uri = None
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == 'typeURI':
                    row_index = cell.row
                    column_index = cell.column
                    filter_uri = sheet.cell(row=row_index + 1, column=column_index).value
        return filter_uri

    @classmethod
    def remove_geo_artefact_excel(cls, workbook):
        for sheet in workbook:
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == 'geometry':
                        sheet.delete_cols(cell.column)

    @classmethod
    def add_choice_list_excel(cls, workbook, instantiated_attributes: list, path_to_subset: Path):
        choice_list_dict = {}
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = SubsetTemplateCreator.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=2):
                for cell in rows:
                    if cell.value.find('[DEPRECATED]') != -1:
                        strip = cell.value.split(' ')
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                strip[1])
                    else:
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                cell.value)

                    if issubclass(dotnotation_attribute.field, KeuzelijstField):
                        name = dotnotation_attribute.field.naam
                        valid_options = [v.invulwaarde for k, v in dotnotation_attribute.field.options.items()
                                         if v.status != 'verwijderd']
                        if (dotnotation_attribute.field.naam not in choice_list_dict):
                            choice_list_dict = cls.add_choice_list_to_sheet(
                                workbook=workbook, name=name,  options=valid_options, choice_list_dict=choice_list_dict)
                        column = choice_list_dict[dotnotation_attribute.field.naam]
                        start_range = f"${column}$2"
                        end_range = f"${column}${len(valid_options) + 1}"
                        data_val = DataValidation(type="list", formula1=f"Keuzelijsten!{start_range}:{end_range}",
                                                  allowBlank=True)
                        sheet.add_data_validation(data_val)
                        data_val.add(f'{get_column_letter(cell.column)}2:{get_column_letter(cell.column)}1000')
                    if issubclass(dotnotation_attribute.field, BooleanField):
                        data_validation = DataValidation(type="list", formula1='"TRUE,FALSE,-"', allow_blank=True)
                        column = cell.column
                        sheet.add_data_validation(data_validation)
                        data_validation.add(f'{get_column_letter(column)}2:{get_column_letter(column)}1000')
                        sheet.add_data_validation(data_validation)

    @classmethod
    def add_mock_data_excel(cls, workbook, rows_of_examples: int):
        for sheet in workbook:
            if sheet == workbook["Keuzelijsten"]:
                break
            if rows_of_examples == 0:
                for rows in sheet.iter_rows(min_row=2, max_row=2):
                    for cell in rows:
                        cell.value = ''

    @classmethod
    def remove_geo_artefact_csv(cls, header, data):
        if 'geometry' in header:
            deletion_index = header.index('geometry')
            header.remove('geometry')
            for d in data:
                d.pop(deletion_index)
        return [header, data]

    @classmethod
    def multiple_csv_template(cls, path_to_template_file_and_extension, path_to_subset, temporary_path,
                              instantiated_attributes, **kwargs):
        file_location = os.path.dirname(path_to_template_file_and_extension)
        tempdir = Path(tempfile.gettempdir()) / 'temp-otlmow'
        logging.debug(file_location)
        file_name = ntpath.basename(path_to_template_file_and_extension)
        split_file_name = file_name.split('.')
        things_in_there = os.listdir(tempdir)
        csv_templates = [x for x in things_in_there if x.startswith(f'{split_file_name[0]}_')]
        for file in csv_templates:
            test_template_loc = Path(os.path.dirname(path_to_template_file_and_extension)) / file
            temp_loc = Path(tempdir) / file
            cls.alter_csv_template(path_to_template_file_and_extension=test_template_loc, temporary_path=temp_loc,
                                   path_to_subset=path_to_subset, **kwargs)

    @classmethod
    def alter_csv_template(cls, path_to_template_file_and_extension, path_to_subset, temporary_path,
                           **kwargs):
        converter = OtlmowConverter()
        instantiated_attributes = converter.from_file_to_objects(file_path=temporary_path,
                                                                    path_to_subset=path_to_subset)
        header = []
        data = []
        delimiter = ';'
        add_geo_artefact = kwargs.get('add_geo_artefact', False)
        add_attribute_info = kwargs.get('add_attribute_info', False)
        highlight_deprecated_attributes = kwargs.get('highlight_deprecated_attributes', False)
        amount_of_examples = kwargs.get('amount_of_examples', 0)
        quote_char = '"'
        with open(temporary_path, 'r+', encoding='utf-8') as csvfile:
            with open(path_to_template_file_and_extension, 'w', encoding='utf-8') as new_file:
                reader = csv.reader(csvfile, delimiter=delimiter, quotechar=quote_char)
                for row_nr, row in enumerate(reader):
                    if row_nr == 0:
                        header = row
                    else:
                        data.append(row)
                if add_geo_artefact is False:
                    [header, data] = cls.remove_geo_artefact_csv(header=header, data=data)
                if add_attribute_info:
                    [info, header] = cls.add_attribute_info_csv(header=header, data=data,
                                                                instantiated_attributes=instantiated_attributes)
                    new_file.write(delimiter.join(info) + '\n')
                data = cls.add_mock_data_csv(header=header, data=data, rows_of_examples=amount_of_examples)
                if highlight_deprecated_attributes:
                    header = cls.highlight_deprecated_attributes_csv(header=header, data=data,
                                                                     instantiated_attributes=instantiated_attributes)
                new_file.write(delimiter.join(header) + '\n')
                for d in data:
                    new_file.write(delimiter.join(d) + '\n')

    @classmethod
    def add_attribute_info_csv(cls, header, data, instantiated_attributes):
        info_data = []
        info_data.extend(header)
        found_uri = []
        dotnotation_module = DotnotationHelper()
        uri_index = cls.find_uri_in_csv(header)
        for d in data:
            if d[uri_index] not in found_uri:
                found_uri.append(d[uri_index])
        for uri in found_uri:
            single_object = next(x for x in instantiated_attributes if x.typeURI == uri)
            for dotnototation_title in info_data:
                if dotnototation_title == 'typeURI':
                    index = info_data.index(dotnototation_title)
                    info_data[index] = 'De URI van het object volgens https://www.w3.org/2001/XMLSchema#anyURI .'
                else:
                    index = info_data.index(dotnototation_title)
                    try:
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(
                            single_object, dotnototation_title)
                    except AttributeError as e:
                        continue
                    info_data[index] = dotnotation_attribute.definition
        return [info_data, header]

    @classmethod
    def add_mock_data_csv(cls, header, data, rows_of_examples):
        if rows_of_examples == 0:
            data = []
        return data

    @classmethod
    def highlight_deprecated_attributes_csv(cls, header, data, instantiated_attributes):
        found_uri = []
        dotnotation_module = DotnotationHelper()
        uri_index = cls.find_uri_in_csv(header)
        for d in data:
            if d[uri_index] not in found_uri:
                found_uri.append(d[uri_index])
        for uri in found_uri:
            single_object = next(x for x in instantiated_attributes if x.typeURI == uri)
            for dotnototation_title in header:
                if dotnototation_title == 'typeURI':
                    continue

                index = header.index(dotnototation_title)
                value = header[index]
                try:
                    is_deprecated = False
                    if dotnototation_title.count('.') == 1:
                        dot_split = dotnototation_title.split('.')
                        attribute = dotnotation_module.get_attribute_by_dotnotation(single_object,
                                                                                    dot_split[0])

                        if len(attribute.deprecated_version) > 0:
                            is_deprecated = True
                    dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_object,
                                                                                            dotnototation_title)
                    if len(dotnotation_attribute.deprecated_version) > 0:
                        is_deprecated = True
                except AttributeError:
                    continue
                if is_deprecated:
                    header[index] = f"[DEPRECATED] {value}"
        return header

    @classmethod
    def find_uri_in_csv(cls, header):
        return header.index('typeURI') if 'typeURI' in header else None

    @classmethod
    def add_choice_list_to_sheet(cls, workbook, name, options, choice_list_dict):
        active_sheet = workbook['Keuzelijsten']
        row_nr = 2
        for rows in active_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=700):
            for cell in rows:
                if cell.value is None:
                    cell.value = name
                    column_nr = cell.column
                    for option in options:
                        active_sheet.cell(row=row_nr, column=column_nr, value=option)
                        row_nr += 1
                    choice_list_dict[name] = get_column_letter(column_nr)
                    break
        return choice_list_dict


if __name__ == '__main__':
    subset_tool = SubsetTemplateCreator()
    subset_location = Path(ROOT_DIR) / 'UnitTests' / 'Subset' / 'Flitspaal_noAgent3.0.db'
    # directory = Path(ROOT_DIR) / 'UnitTests' / 'TestClasses'
    # Slash op het einde toevoegen verandert weinig of niks aan het resultaat
    # directory = os.path.join(directory, '')
    xls_location = Path(ROOT_DIR) / 'UnitTests' / 'Subset' / 'testFileStorage' / 'template_file.csv'
    subset_tool.generate_template_from_subset(path_to_subset=subset_location,
                                              path_to_template_file_and_extension=xls_location, add_attribute_info=True,
                                              highlight_deprecated_attributes=True,
                                              amount_of_examples=5,
                                              generate_choice_list=True,
                                              split_per_type=False)

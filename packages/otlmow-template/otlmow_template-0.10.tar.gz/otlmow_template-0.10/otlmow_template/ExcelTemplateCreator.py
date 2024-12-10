import os
from pathlib import Path
from typing import List

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from otlmow_converter.DotnotationHelper import DotnotationHelper
from otlmow_model.OtlmowModel.BaseClasses.BooleanField import BooleanField
from otlmow_model.OtlmowModel.BaseClasses.KeuzelijstField import KeuzelijstField


class ExcelTemplateCreator:

    @classmethod
    def alter_excel_template(cls, path_to_template_file_and_extension: Path, instantiated_attributes: List,
                             temporary_path: Path, **kwargs):
        generate_choice_list = kwargs.get('generate_choice_list', False)
        add_geo_artefact = kwargs.get('add_geo_artefact', False)
        add_attribute_info = kwargs.get('add_attribute_info', False)
        highlight_deprecated_attributes = kwargs.get('highlight_deprecated_attributes', False)
        amount_of_examples = kwargs.get('amount_of_examples', 0)
        wb = load_workbook(temporary_path)
        wb.create_sheet('Keuzelijsten')
        # Volgorde is belangrijk! Eerst rijen verwijderen indien nodig dan choice list toevoegen,
        # staat namelijk vast op de kolom en niet het attribuut in die kolom
        if add_geo_artefact is False:
            cls.remove_geo_artefact_excel(workbook=wb)
        if generate_choice_list:
            cls.add_choice_list_excel(workbook=wb, instantiated_attributes=instantiated_attributes)
        cls.add_mock_data_excel(workbook=wb, rows_of_examples=amount_of_examples)
        if highlight_deprecated_attributes:
            cls.check_for_deprecated_attributes_excel(workbook=wb, instantiated_attributes=instantiated_attributes)
        if add_attribute_info:
            cls.add_attribute_info_excel(workbook=wb, instantiated_attributes=instantiated_attributes)
        cls.design_workbook_excel(workbook=wb)
        wb.save(path_to_template_file_and_extension)
        file_location = os.path.dirname(temporary_path)
        [f.unlink() for f in Path(file_location).glob("*") if f.is_file()]

    @classmethod
    def add_attribute_info_excel(cls, workbook, instantiated_attributes: List):
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = cls.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            sheet.insert_rows(1)
            for rows in sheet.iter_rows(min_row=2, max_row=2, min_col=1):
                for cell in rows:
                    if cell.value == 'typeURI':
                        value = 'De URI van het object volgens https://www.w3.org/2001/XMLSchema#anyURI .'
                    elif cell.value.find('[DEPRECATED]') != -1:
                        strip = cell.value.split(' ')
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                strip[1])
                        value = dotnotation_attribute.definition
                    else:
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                cell.value)
                        value = dotnotation_attribute.definition

                    sheet.cell(row=1, column=cell.column, value=value)
                    sheet.cell(row=1, column=cell.column).fill = PatternFill(start_color="808080", end_color="808080",
                                                                             fill_type="solid")

    @classmethod
    def check_for_deprecated_attributes_excel(cls, workbook, instantiated_attributes: List):
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = cls.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=2):
                for cell in rows:
                    is_deprecated = False
                    if cell.value.count('.') == 1:
                        dot_split = cell.value.split('.')
                        attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                    dot_split[0])

                        if len(attribute.deprecated_version) > 0:
                            is_deprecated = True
                    dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                            cell.value)
                    if len(dotnotation_attribute.deprecated_version) > 0:
                        is_deprecated = True

                    if is_deprecated:
                        cell.value = '[DEPRECATED] ' + cell.value

    @classmethod
    def find_uri_in_sheet(cls, sheet):
        filter_uri = None
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == 'typeURI':
                    row_index = cell.row
                    column_index = cell.column
                    filter_uri = sheet.cell(row=row_index + 1, column=column_index).value
        return filter_uri

    @classmethod
    def remove_geo_artefact_excel(cls, workbook):
        for sheet in workbook:
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value == 'geometry':
                        sheet.delete_cols(cell.column)

    @classmethod
    def add_choice_list_excel(cls, workbook, instantiated_attributes: List):
        choice_list_dict = {}
        dotnotation_module = DotnotationHelper()
        for sheet in workbook:
            if sheet == workbook['Keuzelijsten']:
                break
            filter_uri = cls.find_uri_in_sheet(sheet)
            single_attribute = next(x for x in instantiated_attributes if x.typeURI == filter_uri)
            for rows in sheet.iter_rows(min_row=1, max_row=1, min_col=2):
                for cell in rows:
                    if cell.value.find('[DEPRECATED]') != -1:
                        strip = cell.value.split(' ')
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                strip[1])
                    else:
                        dotnotation_attribute = dotnotation_module.get_attribute_by_dotnotation(single_attribute,
                                                                                                cell.value)

                    if issubclass(dotnotation_attribute.field, KeuzelijstField):
                        name = dotnotation_attribute.field.naam
                        valid_options = [v.invulwaarde for k, v in dotnotation_attribute.field.options.items()
                                         if v.status != 'verwijderd']
                        column = cls.return_column_letter_of_choice_list(name=name, choice_list_dict=choice_list_dict,
                                                                         options=valid_options, workbook=workbook)
                        option_list = []
                        for option in valid_options:
                            option_list.append(option)
                        start_range = f"${column}$2"
                        end_range = f"${column}${len(valid_options) + 1}"
                        data_val = DataValidation(type="list", formula1=f"Keuzelijsten!{start_range}:{end_range}",
                                                  allowBlank=True)
                        sheet.add_data_validation(data_val)
                        data_val.add(f'{get_column_letter(cell.column)}2:{get_column_letter(cell.column)}1000')
                    if issubclass(dotnotation_attribute.field, BooleanField):
                        data_validation = DataValidation(type="list", formula1='"TRUE,FALSE,-"', allow_blank=True)
                        column = cell.column
                        sheet.add_data_validation(data_validation)
                        data_validation.add(f'{get_column_letter(column)}2:{get_column_letter(column)}1000')
                        sheet.add_data_validation(data_validation)

    @classmethod
    def add_mock_data_excel(cls, workbook, rows_of_examples: int):
        for sheet in workbook:
            if sheet == workbook["Keuzelijsten"]:
                break
            if rows_of_examples == 0:
                for rows in sheet.iter_rows(min_row=2, max_row=2):
                    for cell in rows:
                        cell.value = ''

    @classmethod
    def return_column_letter_of_choice_list(cls, name, choice_list_dict, options, workbook):
        if name in choice_list_dict:
            column = choice_list_dict[name]
        else:
            choice_list_dict = cls.add_choice_list_to_sheet(workbook=workbook, name=name,
                                                            options=options,
                                                            choice_list_dict=choice_list_dict)
            column = choice_list_dict[name]
        return column

    @classmethod
    def add_choice_list_to_sheet(cls, workbook, name, options, choice_list_dict):
        active_sheet = workbook['Keuzelijsten']
        row_nr = 2
        for rows in active_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=700):
            for cell in rows:
                if cell.value is None:
                    cell.value = name
                    column_nr = cell.column
                    for option in options:
                        active_sheet.cell(row=row_nr, column=column_nr, value=option)
                        row_nr += 1
                    choice_list_dict[name] = get_column_letter(column_nr)
                    break
        return choice_list_dict

    @classmethod
    def design_workbook_excel(cls, workbook):
        for sheet in workbook:
            dim_holder = DimensionHolder(worksheet=sheet)
            for col in range(sheet.min_column, sheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)
            sheet.column_dimensions = dim_holder
